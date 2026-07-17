from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File, Form
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.staticfiles import StaticFiles
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from starlette.responses import StreamingResponse, FileResponse
import os
import logging
import io
import csv
import json
import re
import shutil
import asyncio
import resend
import secrets
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import jwt
from passlib.context import CryptContext
import pandas as pd
from openpyxl import load_workbook

# PDF Generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm, mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Resend Configuration
RESEND_API_KEY = os.environ.get('RESEND_API_KEY')
SENDER_EMAIL = os.environ.get('SENDER_EMAIL', 'onboarding@resend.dev')
ADMIN_EMAIL = os.environ.get('ADMIN_EMAIL', '')

if RESEND_API_KEY:
    resend.api_key = RESEND_API_KEY

# Create uploads directory
UPLOADS_DIR = ROOT_DIR / "uploads"
UPLOADS_DIR.mkdir(exist_ok=True)
(UPLOADS_DIR / "equipments").mkdir(exist_ok=True)
(UPLOADS_DIR / "inspections").mkdir(exist_ok=True)
(UPLOADS_DIR / "subequipments").mkdir(exist_ok=True)
(UPLOADS_DIR / "spareparts").mkdir(exist_ok=True)
(UPLOADS_DIR / "workorders").mkdir(exist_ok=True)
(UPLOADS_DIR / "interventions").mkdir(exist_ok=True)
(UPLOADS_DIR / "contractors").mkdir(exist_ok=True)
(UPLOADS_DIR / "gas_cylinders").mkdir(exist_ok=True)
(UPLOADS_DIR / "contracts").mkdir(exist_ok=True)
(UPLOADS_DIR / "documents").mkdir(exist_ok=True)
(UPLOADS_DIR / "reports").mkdir(exist_ok=True)
(UPLOADS_DIR / "imports").mkdir(exist_ok=True)

# Currency conversion rate (XPF to EUR)
XPF_TO_EUR = 0.00838  # 1 XPF = 0.00838 EUR (taux fixe)

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Configuration
JWT_SECRET = os.environ['JWT_SECRET']
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 24

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()

app = FastAPI(title="HyperbareManager API")
api_router = APIRouter(prefix="/api")

# ==================== EMAIL SERVICE ====================

async def send_email(to_email: str, subject: str, html_content: str) -> bool:
    """Send email using Resend API"""
    if not RESEND_API_KEY:
        logging.warning("RESEND_API_KEY not configured - email not sent")
        return False
    
    try:
        params = {
            "from": SENDER_EMAIL,
            "to": [to_email],
            "subject": subject,
            "html": html_content
        }
        await asyncio.to_thread(resend.Emails.send, params)
        logging.info(f"Email sent to {to_email}: {subject}")
        return True
    except Exception as e:
        logging.error(f"Failed to send email to {to_email}: {str(e)}")
        return False

def email_template(title: str, content: str, footer: str = "") -> str:
    """Generate HTML email template"""
    return f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
    </head>
    <body style="margin: 0; padding: 0; font-family: Arial, sans-serif; background-color: #f4f4f4;">
        <table width="100%" cellpadding="0" cellspacing="0" style="max-width: 600px; margin: 0 auto; background-color: #ffffff;">
            <tr>
                <td style="background-color: #005F73; padding: 20px; text-align: center;">
                    <h1 style="color: #ffffff; margin: 0; font-size: 24px;">HyperbareManager</h1>
                </td>
            </tr>
            <tr>
                <td style="padding: 30px;">
                    <h2 style="color: #005F73; margin-top: 0;">{title}</h2>
                    {content}
                </td>
            </tr>
            <tr>
                <td style="background-color: #f8f9fa; padding: 20px; text-align: center; color: #6c757d; font-size: 12px;">
                    {footer if footer else "Cet email a été envoyé automatiquement par HyperbareManager."}
                </td>
            </tr>
        </table>
    </body>
    </html>
    """

async def send_welcome_email(user_email: str, user_name: str, password: str):
    """Send welcome email to newly created user"""
    content = f"""
    <p>Bonjour <strong>{user_name}</strong>,</p>
    <p>Votre compte HyperbareManager a été créé avec succès.</p>
    <table style="background-color: #f8f9fa; padding: 15px; border-radius: 5px; width: 100%;">
        <tr>
            <td style="padding: 10px;">
                <strong>Email :</strong> {user_email}<br>
                <strong>Mot de passe :</strong> {password}
            </td>
        </tr>
    </table>
    <p style="margin-top: 20px;">Nous vous recommandons de changer votre mot de passe après votre première connexion.</p>
    <p>Connectez-vous dès maintenant pour commencer à utiliser l'application.</p>
    """
    await send_email(user_email, "Bienvenue sur HyperbareManager", email_template("Bienvenue !", content))

async def send_access_approved_email(user_email: str, user_name: str):
    """Send email when access request is approved"""
    content = f"""
    <p>Bonjour <strong>{user_name}</strong>,</p>
    <p style="color: #28a745;"><strong>Bonne nouvelle !</strong> Votre demande d'accès à HyperbareManager a été approuvée.</p>
    <p>Vous pouvez maintenant vous connecter avec vos identifiants.</p>
    """
    await send_email(user_email, "Accès approuvé - HyperbareManager", email_template("Accès Approuvé ✓", content))

async def send_access_rejected_email(user_email: str, user_name: str):
    """Send email when access request is rejected"""
    content = f"""
    <p>Bonjour <strong>{user_name}</strong>,</p>
    <p style="color: #dc3545;">Nous avons le regret de vous informer que votre demande d'accès à HyperbareManager a été refusée.</p>
    <p>Si vous pensez qu'il s'agit d'une erreur, veuillez contacter l'administrateur.</p>
    """
    await send_email(user_email, "Demande d'accès refusée - HyperbareManager", email_template("Demande Refusée", content))

async def send_password_reset_request_email(admin_email: str, requester_name: str, requester_email: str):
    """Notify an admin that a user requested a password reset."""
    content = f"""
    <p>Bonjour,</p>
    <p>L'utilisateur <strong>{requester_name}</strong> ({requester_email}) a demandé la réinitialisation de son mot de passe.</p>
    <p>Connectez-vous à HyperbareManager, ouvrez la page <strong>Utilisateurs</strong>, puis cliquez sur
    <strong>« Envoyer un mot de passe temporaire »</strong> pour lui transmettre de nouveaux identifiants.</p>
    """
    return await send_email(admin_email, "🔑 Demande de réinitialisation de mot de passe", email_template("Réinitialisation demandée", content))

async def send_temp_password_email(user_email: str, user_name: str, temp_password: str):
    """Send a temporary password to a user."""
    content = f"""
    <p>Bonjour <strong>{user_name}</strong>,</p>
    <p>Un mot de passe temporaire vous a été attribué par l'administrateur.</p>
    <table style="background-color: #f8f9fa; padding: 15px; border-radius: 5px; width: 100%;">
        <tr>
            <td style="padding: 10px;">
                <strong>Email :</strong> {user_email}<br>
                <strong>Mot de passe temporaire :</strong> <span style="font-family: monospace; font-size: 16px;">{temp_password}</span>
            </td>
        </tr>
    </table>
    <p style="margin-top: 20px; color: #AE2012;">Pour des raisons de sécurité, vous devrez définir un nouveau mot de passe dès votre prochaine connexion.</p>
    """
    return await send_email(user_email, "🔑 Mot de passe temporaire - HyperbareManager", email_template("Mot de passe temporaire", content))

async def send_maintenance_reminder_email(to_email: str, maintenance_title: str, equipment_ref: str, date_planifiee: str, days_left: int):
    """Send maintenance reminder email"""
    urgency_color = "#dc3545" if days_left <= 7 else "#ffc107" if days_left <= 14 else "#17a2b8"
    content = f"""
    <p>Une maintenance préventive est prévue prochainement :</p>
    <table style="background-color: #f8f9fa; padding: 15px; border-radius: 5px; width: 100%; border-left: 4px solid {urgency_color};">
        <tr>
            <td style="padding: 10px;">
                <strong>Titre :</strong> {maintenance_title}<br>
                <strong>Équipement :</strong> {equipment_ref}<br>
                <strong>Date planifiée :</strong> {date_planifiee}<br>
                <strong style="color: {urgency_color};">Dans {days_left} jour(s)</strong>
            </td>
        </tr>
    </table>
    <p style="margin-top: 20px;">Pensez à planifier cette intervention.</p>
    """
    await send_email(to_email, f"⏰ Rappel maintenance : {maintenance_title}", email_template("Maintenance à Venir", content))

async def send_maintenance_overdue_email(to_email: str, maintenance_title: str, equipment_ref: str, date_planifiee: str, days_overdue: int):
    """Send overdue maintenance alert email"""
    content = f"""
    <p style="color: #dc3545;"><strong>⚠️ ALERTE : Maintenance en retard !</strong></p>
    <table style="background-color: #fff5f5; padding: 15px; border-radius: 5px; width: 100%; border-left: 4px solid #dc3545;">
        <tr>
            <td style="padding: 10px;">
                <strong>Titre :</strong> {maintenance_title}<br>
                <strong>Équipement :</strong> {equipment_ref}<br>
                <strong>Date planifiée :</strong> {date_planifiee}<br>
                <strong style="color: #dc3545;">En retard de {days_overdue} jour(s)</strong>
            </td>
        </tr>
    </table>
    <p style="margin-top: 20px;">Veuillez effectuer cette maintenance au plus vite.</p>
    """
    await send_email(to_email, f"🚨 URGENT - Maintenance en retard : {maintenance_title}", email_template("Maintenance en Retard", content))

async def send_low_stock_email(to_email: str, part_name: str, part_ref: str, current_stock: int, minimum_stock: int):
    """Send low stock alert email"""
    content = f"""
    <p style="color: #ffc107;"><strong>⚠️ Alerte stock bas</strong></p>
    <table style="background-color: #fff9e6; padding: 15px; border-radius: 5px; width: 100%; border-left: 4px solid #ffc107;">
        <tr>
            <td style="padding: 10px;">
                <strong>Pièce :</strong> {part_name}<br>
                <strong>Référence :</strong> {part_ref}<br>
                <strong>Stock actuel :</strong> <span style="color: #dc3545;">{current_stock}</span><br>
                <strong>Seuil minimum :</strong> {minimum_stock}
            </td>
        </tr>
    </table>
    <p style="margin-top: 20px;">Pensez à réapprovisionner cette pièce.</p>
    """
    await send_email(to_email, f"📦 Stock bas : {part_name}", email_template("Alerte Stock Bas", content))

async def send_hour_counter_alert_email(to_email: str, equipment_ref: str, current_hours: float, threshold_hours: float, maintenance_title: str):
    """Send compressor hour counter alert email"""
    content = f"""
    <p style="color: #dc3545;"><strong>⚠️ Seuil compteur horaire atteint !</strong></p>
    <table style="background-color: #fff5f5; padding: 15px; border-radius: 5px; width: 100%; border-left: 4px solid #dc3545;">
        <tr>
            <td style="padding: 10px;">
                <strong>Équipement :</strong> {equipment_ref}<br>
                <strong>Compteur actuel :</strong> {current_hours:,.0f} h<br>
                <strong>Seuil de déclenchement :</strong> {threshold_hours:,.0f} h<br>
                <strong>Maintenance requise :</strong> {maintenance_title}
            </td>
        </tr>
    </table>
    <p style="margin-top: 20px;">Une maintenance basée sur le compteur horaire doit être effectuée.</p>
    """
    await send_email(to_email, f"🔧 Compteur horaire : {equipment_ref} - Maintenance requise", email_template("Alerte Compteur Horaire", content))

# ==================== MODELS ====================

# User Models
ROLES = ["admin", "technicien", "invite"]

class UserBase(BaseModel):
    email: EmailStr
    nom: str
    prenom: str
    role: str = Field(default="invite", description="admin, technicien, or invite")

class UserCreate(UserBase):
    password: str

class User(UserBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    is_active: bool = False  # Requires admin approval
    is_approved: bool = False  # Approval status

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: dict

# Caisson Model
class CaissonBase(BaseModel):
    identifiant: str
    modele: str
    fabricant: str
    date_mise_en_service: str
    pression_maximale: float
    normes_applicables: List[str] = []
    description: Optional[str] = None

class CaissonCreate(CaissonBase):
    pass

class Caisson(CaissonBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Equipment Type Model (Dynamic types)
class EquipmentTypeBase(BaseModel):
    nom: str
    description: Optional[str] = None
    icon: Optional[str] = None

class EquipmentTypeCreate(EquipmentTypeBase):
    pass

class EquipmentType(EquipmentTypeBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Equipment Model
class EquipmentBase(BaseModel):
    type: str  # porte, joint, soupape, compresseur, capteur, systeme_securite
    reference: str
    numero_serie: str
    criticite: str = Field(default="normale", description="critique, haute, normale, basse")
    statut: str = Field(default="en_service", description="en_service, maintenance, hors_service, reforme")
    caisson_id: str
    description: Optional[str] = None
    date_installation: Optional[str] = None
    date_reforme: Optional[str] = None  # Date de réforme (si statut = reforme)
    motif_reforme: Optional[str] = None  # Motif de réforme
    technicien_reforme: Optional[str] = None  # Technicien responsable de la réforme
    gas_cylinder_id: Optional[str] = None  # Bouteille de gaz associée (ex: extincteur, ARI)
    photos: List[str] = []  # Liste des URLs des photos
    documents: List[dict] = []  # Liste des documents PDF [{filename, url, uploaded_at}]
    compteur_horaire: Optional[float] = None  # Compteur horaire pour les compresseurs (en heures)
    historique_compteur: List[dict] = []  # Historique des relevés [{date, valeur, technicien}]

class EquipmentCreate(EquipmentBase):
    pass

class Equipment(EquipmentBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    historique_statut: List[dict] = []  # Journal des changements de statut [{date, ancien, nouveau, motif, utilisateur}]

# Sub-Equipment Model (Sous-équipement)
class SubEquipmentBase(BaseModel):
    nom: str
    reference: str
    numero_serie: Optional[str] = None
    parent_equipment_id: str  # Lien vers l'équipement parent
    description: Optional[str] = None
    date_installation: Optional[str] = None
    statut: str = Field(default="en_service", description="en_service, maintenance, hors_service")
    photos: List[str] = []
    documents: List[dict] = []

class SubEquipmentCreate(SubEquipmentBase):
    pass

class SubEquipment(SubEquipmentBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Work Order Model
class WorkOrderBase(BaseModel):
    titre: str
    description: str
    type_maintenance: str  # preventive, corrective
    priorite: str = Field(default="normale", description="urgente, haute, normale, basse")
    statut: str = Field(default="planifiee", description="planifiee, en_cours, terminee, annulee")
    caisson_id: Optional[str] = None
    equipment_id: Optional[str] = None
    date_planifiee: str
    periodicite_jours: Optional[int] = None
    # Pour les compresseurs - maintenance basée sur le compteur horaire
    periodicite_heures: Optional[int] = None  # Périodicité en heures de fonctionnement
    compteur_declenchement: Optional[float] = None  # Compteur horaire au moment où la maintenance doit être faite
    technicien_assigne: Optional[str] = None
    photos: List[str] = []
    documents: List[dict] = []

class WorkOrderCreate(WorkOrderBase):
    pass

class WorkOrder(WorkOrderBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Intervention Model
class InterventionBase(BaseModel):
    work_order_id: Optional[str] = None  # Pour maintenance curative (ordre de travail) — désormais optionnel
    maintenance_preventive_id: Optional[str] = None  # Pour maintenance préventive (inspection)
    type_intervention: str = Field(default="curative", description="curative ou preventive")
    titre: Optional[str] = None  # Motif / désignation de l'intervention curative
    date_intervention: str
    technicien: str
    actions_realisees: str
    observations: Optional[str] = None
    pieces_utilisees: List[dict] = []  # [{spare_part_id, quantite}]
    duree_minutes: Optional[int] = None
    compteur_horaire: Optional[float] = None  # Compteur horaire au moment de l'intervention (pour compresseurs)
    equipment_id: Optional[str] = None  # Équipement concerné
    sous_equipement_id: Optional[str] = None  # Sous-équipement concerné (optionnel)
    documents: List[dict] = []  # PV / documents PDF [{filename, url, uploaded_at}]

class InterventionCreate(InterventionBase):
    pass

class Intervention(InterventionBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Formation (Training slot) Model
class FormationBase(BaseModel):
    nom: str
    technicien: str  # Nom affiché du technicien
    technicien_id: Optional[str] = None  # ID utilisateur (pour filtrage par rôle)
    date_debut: str  # YYYY-MM-DD
    date_fin: str    # YYYY-MM-DD
    description: Optional[str] = None

class FormationCreate(FormationBase):
    pass

class Formation(FormationBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Inspection (Contrôle réglementaire) Model
PERIODICITES = {
    "hebdomadaire": 7,
    "mensuel": 30,
    "trimestriel": 90,
    "semestriel": 180,
    "annuel": 365,
    "biannuel": 730,
    "triennal": 1095,
    "quinquennal": 1825,
    "decennal": 3650,
    "journalier": 1,
    "quotidien": 1,
}

def _norm_periodicite(value: str) -> str:
    """Normalise une périodicité saisie (accents, casse) vers une clé valide de PERIODICITES."""
    if not value:
        return "annuel"
    import unicodedata
    s = unicodedata.normalize("NFKD", str(value).strip().lower())
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s if s in PERIODICITES else "annuel"

class InspectionBase(BaseModel):
    titre: str
    type_controle: str
    periodicite: str = Field(default="annuel", description="hebdomadaire, mensuel, trimestriel, semestriel, annuel, biannuel, triennal, quinquennal, decennal")
    caisson_id: Optional[str] = None
    equipment_id: Optional[str] = None
    date_realisation: Optional[str] = None
    date_validite: Optional[str] = None  # Calculée automatiquement
    organisme_certificateur: Optional[str] = None
    resultat: Optional[str] = None
    observations: Optional[str] = None
    procedure_documents: List[dict] = []  # Liste des procédures PDF [{filename, url, uploaded_at}]
    historique_controles: List[dict] = []  # Historique des réalisations passées (traçabilité)

class InspectionCreate(InspectionBase):
    pass

class Inspection(InspectionBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Spare Part Model
class SparePartBase(BaseModel):
    nom: str
    reference_fabricant: str
    equipment_type: str  # type d'équipement concerné
    quantite_stock: int = 0
    seuil_minimum: int = 1
    emplacement: Optional[str] = None
    fournisseur: Optional[str] = None
    prix_unitaire: Optional[float] = None
    photos: List[str] = []
    documents: List[dict] = []

class SparePartCreate(SparePartBase):
    pass

class SparePart(SparePartBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SparePartUpdate(BaseModel):
    quantite_stock: Optional[int] = None
    seuil_minimum: Optional[int] = None
    emplacement: Optional[str] = None
    fournisseur: Optional[str] = None
    prix_unitaire: Optional[float] = None

# ==================== NEW MODELS FOR EXTENDED FEATURES ====================

# Prestataire/Fournisseur Model
class ContractorBase(BaseModel):
    nom: str
    type: str = Field(default="prestataire", description="prestataire, fournisseur, organisme_controle")
    specialite: Optional[str] = None  # Maintenance compresseurs, Métrologie, etc.
    contact_nom: Optional[str] = None
    contact_email: Optional[str] = None
    contact_telephone: Optional[str] = None
    adresse: Optional[str] = None
    siret: Optional[str] = None
    notes: Optional[str] = None
    documents: List[dict] = []  # Contrats, certifications, etc.

class ContractorCreate(ContractorBase):
    pass

class Contractor(ContractorBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Bouteille de Gaz Model
GAS_TYPES = ["O2", "air_medicale", "heliox", "nitrox"]

class GasCylinderBase(BaseModel):
    numero_bouteille: str
    type_gaz: str = Field(description="O2, air_medicale, heliox, nitrox")
    volume: str = Field(default="B50", description="B5, B50, etc.")
    pression_service: Optional[float] = None  # bars
    fournisseur_id: Optional[str] = None
    localisation: Optional[str] = None
    date_remplissage: Optional[str] = None
    date_expiration_gaz: Optional[str] = None  # Péremption du gaz
    date_epreuve: Optional[str] = None  # Dernière épreuve hydraulique
    date_prochaine_epreuve: Optional[str] = None  # Prochaine requalification (5 ans)
    statut: str = Field(default="pleine", description="pleine, en_cours, vide, hors_service")
    observations: Optional[str] = None
    agent_responsable: Optional[str] = None
    documents: List[dict] = []

class GasCylinderCreate(GasCylinderBase):
    pass

class GasCylinder(GasCylinderBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    historique_remplissage: List[dict] = []  # [{date, agent, pression, observations}]

# Contrat de Maintenance Model
class MaintenanceContractBase(BaseModel):
    numero_contrat: str
    titre: str
    contractor_id: str  # Lien vers prestataire
    type_contrat: str = Field(default="maintenance", description="maintenance, controle, fourniture")
    date_debut: str
    date_fin: str
    montant_annuel: Optional[float] = None
    devise: str = Field(default="XPF")
    periodicite_facturation: Optional[str] = None  # mensuel, trimestriel, annuel
    prestations_incluses: List[str] = []  # Liste des prestations couvertes
    equipements_couverts: List[str] = []  # IDs des équipements
    conditions_particulieres: Optional[str] = None
    statut: str = Field(default="actif", description="actif, suspendu, expire, resilie")
    documents: List[dict] = []  # Contrat signé, avenants, etc.

class MaintenanceContractCreate(MaintenanceContractBase):
    pass

class MaintenanceContract(MaintenanceContractBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Document Model (Gestion documentaire)
class DocumentBase(BaseModel):
    titre: str
    type_document: str = Field(description="notice, rapport, certificat, plan, procedure, pv_controle, autre")
    categorie: Optional[str] = None  # Equipement, Sécurité, Réglementaire, etc.
    description: Optional[str] = None
    equipment_id: Optional[str] = None
    contractor_id: Optional[str] = None
    date_document: Optional[str] = None
    date_validite: Optional[str] = None  # Pour les certificats
    version: Optional[str] = None
    fichier_url: Optional[str] = None
    fichier_nom: Optional[str] = None
    tags: List[str] = []

class DocumentCreate(DocumentBase):
    pass

class Document(DocumentBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    uploaded_by: Optional[str] = None

# Budget Item Model
class BudgetItemBase(BaseModel):
    annee: int
    categorie: str = Field(description="maintenance_preventive, maintenance_corrective, controle_reglementaire, pieces_detachees, consommables, prestation_externe, renouvellement")
    designation: str
    description: Optional[str] = None
    equipment_id: Optional[str] = None
    contractor_id: Optional[str] = None
    periodicite: Optional[str] = None  # 1 mois, 1 an, 1000 heures, etc.
    montant_prevu_xpf: float = 0
    montant_prevu_eur: Optional[float] = None  # Calculé automatiquement
    montant_realise_xpf: Optional[float] = None
    montant_realise_eur: Optional[float] = None
    date_prevue: Optional[str] = None
    date_realisee: Optional[str] = None
    statut: str = Field(default="prevu", description="prevu, en_cours, realise, annule")
    notes: Optional[str] = None

class BudgetItemCreate(BudgetItemBase):
    pass

class BudgetItem(BudgetItemBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Modèle de PV de contrôle
class ReportTemplateBase(BaseModel):
    nom: str
    type_controle: str = Field(description="analyse_air, controle_mensuel, controle_trimestriel, controle_semestriel, controle_annuel, etalonnage_manometre, etalonnage_soupape")
    description: Optional[str] = None
    champs: List[dict] = []  # [{nom, type, obligatoire, valeur_defaut, options}]
    normes_reference: List[str] = []  # Normes applicables
    criteres_conformite: List[dict] = []  # [{parametre, valeur_min, valeur_max, unite}]
    modele_actif: bool = True

class ReportTemplateCreate(ReportTemplateBase):
    pass

class ReportTemplate(ReportTemplateBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# PV de Contrôle généré
class ControlReportBase(BaseModel):
    template_id: str
    equipment_id: Optional[str] = None
    inspection_id: Optional[str] = None
    numero_pv: str
    date_controle: str
    controleur: str
    organisme: Optional[str] = None
    valeurs: dict = {}  # Valeurs saisies pour chaque champ du template
    resultat: str = Field(default="conforme", description="conforme, non_conforme, avec_reserves")
    observations: Optional[str] = None
    documents: List[dict] = []  # PV signé, photos, etc.
    validite_jusqua: Optional[str] = None

class ControlReportCreate(ControlReportBase):
    pass

class ControlReport(ControlReportBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# ==================== AUTH HELPERS ====================

def hash_password(password: str) -> str:
    return pwd_context.hash(password)

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Token invalide")
        user = await db.users.find_one({"id": user_id}, {"_id": 0})
        if user is None:
            raise HTTPException(status_code=401, detail="Utilisateur non trouvé")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expiré")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token invalide")

async def require_admin(current_user: dict = Depends(get_current_user)):
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs")
    return current_user

async def require_technicien_or_admin(current_user: dict = Depends(get_current_user)):
    """Allow admin and technicien roles - can create/modify but technicien cannot delete"""
    if current_user.get("role") not in ["admin", "technicien"]:
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs et techniciens")
    return current_user

async def require_active_user(current_user: dict = Depends(get_current_user)):
    """Ensure user is active and approved"""
    if not current_user.get("is_active") or not current_user.get("is_approved"):
        raise HTTPException(status_code=403, detail="Compte non activé. Veuillez contacter l'administrateur.")
    return current_user

def can_delete(user: dict) -> bool:
    """Only admin can delete"""
    return user.get("role") == "admin"

def can_modify(user: dict) -> bool:
    """Admin and technicien can modify"""
    return user.get("role") in ["admin", "technicien"]

def can_export(user: dict) -> bool:
    """Only admin and technicien can export"""
    return user.get("role") in ["admin", "technicien"]

# ==================== AUTH ROUTES ====================

@api_router.post("/auth/register", response_model=dict)
async def register(user_data: UserCreate):
    existing = await db.users.find_one({"email": user_data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email déjà utilisé")
    
    user_dict = user_data.model_dump()
    password = user_dict.pop("password")
    
    # Check if first user - make them admin and auto-approve
    user_count = await db.users.count_documents({})
    if user_count == 0:
        user_dict["role"] = "admin"
        user_dict["is_active"] = True
        user_dict["is_approved"] = True
    else:
        user_dict["role"] = "invite"  # New users start as invite
        user_dict["is_active"] = False
        user_dict["is_approved"] = False
    
    user_obj = User(**user_dict)
    
    doc = user_obj.model_dump()
    doc["password_hash"] = hash_password(password)
    doc["created_at"] = doc["created_at"].isoformat()
    
    await db.users.insert_one(doc)
    
    # If not approved, return message instead of token
    if not user_obj.is_approved:
        return {
            "message": "Inscription réussie. Votre compte est en attente d'approbation par l'administrateur.",
            "pending_approval": True,
            "user": {"id": user_obj.id, "email": user_obj.email, "nom": user_obj.nom, "prenom": user_obj.prenom}
        }
    
    token = create_access_token({"sub": user_obj.id, "email": user_obj.email, "role": user_obj.role})
    
    return {
        "access_token": token,
        "token_type": "bearer",
        "user": {"id": user_obj.id, "email": user_obj.email, "nom": user_obj.nom, "prenom": user_obj.prenom, "role": user_obj.role}
    }

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(credentials: UserLogin):
    user = await db.users.find_one({"email": credentials.email}, {"_id": 0})
    if not user or not verify_password(credentials.password, user.get("password_hash", "")):
        raise HTTPException(status_code=401, detail="Email ou mot de passe incorrect")
    
    # Check if user is approved and active
    if not user.get("is_approved", False):
        raise HTTPException(status_code=403, detail="Votre compte est en attente d'approbation par l'administrateur")
    
    if not user.get("is_active", False):
        raise HTTPException(status_code=403, detail="Votre compte a été suspendu. Contactez l'administrateur.")
    
    token = create_access_token({"sub": user["id"], "email": user["email"], "role": user["role"]})
    
    return TokenResponse(
        access_token=token,
        user={"id": user["id"], "email": user["email"], "nom": user["nom"], "prenom": user["prenom"], "role": user["role"], "must_change_password": user.get("must_change_password", False)}
    )

@api_router.get("/auth/me")
async def get_me(current_user: dict = Depends(get_current_user)):
    return current_user

class ForgotPasswordRequest(BaseModel):
    email: str

def generate_temp_password(length: int = 10) -> str:
    """Generate a readable secure temporary password (no ambiguous chars)."""
    alphabet = "ABCDEFGHJKLMNPQRSTUVWXYZabcdefghijkmnpqrstuvwxyz23456789"
    return "".join(secrets.choice(alphabet) for _ in range(length))

@api_router.post("/auth/forgot-password")
async def forgot_password(data: ForgotPasswordRequest):
    """Public: a user requests a password reset. Admins are notified by email.
    Returns a generic message to avoid leaking whether the account exists."""
    email = (data.email or "").strip().lower()
    generic = {"message": "Si un compte est associé à cet email, les administrateurs ont été notifiés."}
    if not email:
        return generic
    user = await db.users.find_one({"email": email})
    if not user:
        return generic

    now = datetime.now(timezone.utc).isoformat()
    existing = await db.password_reset_requests.find_one({"user_id": user["id"], "statut": "pending"})
    if existing:
        await db.password_reset_requests.update_one({"id": existing["id"]}, {"$set": {"created_at": now}})
    else:
        await db.password_reset_requests.insert_one({
            "id": str(uuid.uuid4()),
            "user_id": user["id"],
            "email": user["email"],
            "nom": user.get("nom", ""),
            "prenom": user.get("prenom", ""),
            "statut": "pending",
            "created_at": now,
        })

    requester_name = f"{user.get('prenom', '')} {user.get('nom', '')}".strip() or user["email"]
    admins = await db.users.find({"role": "admin", "is_active": True}, {"_id": 0, "email": 1}).to_list(100)
    for a in admins:
        if a.get("email"):
            await send_password_reset_request_email(a["email"], requester_name, user["email"])
    return generic

# ==================== USERS ROUTES (Admin only) ====================

@api_router.get("/users", response_model=List[dict])
async def get_users(admin: dict = Depends(require_admin)):
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(1000)
    return users

@api_router.get("/users/pending", response_model=List[dict])
async def get_pending_users(admin: dict = Depends(require_admin)):
    """Get users pending approval"""
    users = await db.users.find({"is_approved": False}, {"_id": 0, "password_hash": 0}).to_list(1000)
    return users

@api_router.get("/users/technicians", response_model=List[dict])
async def get_technicians(current_user: dict = Depends(get_current_user)):
    """Get active users for technician dropdown (admins excluded)."""
    users = await db.users.find(
        {"is_active": True, "is_approved": True, "role": {"$ne": "admin"}},
        {"_id": 0, "password_hash": 0}
    ).to_list(1000)
    return users

# Admin create user
class AdminUserCreate(BaseModel):
    email: str
    nom: str
    prenom: str
    password: str
    role: str = "technicien"

@api_router.post("/users/create")
async def admin_create_user(user_data: AdminUserCreate, admin: dict = Depends(require_admin)):
    """Admin creates a new user directly (pre-approved)"""
    # Check if email already exists
    existing = await db.users.find_one({"email": user_data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Cet email est déjà utilisé")
    
    if user_data.role not in ROLES:
        raise HTTPException(status_code=400, detail=f"Rôle invalide. Choix: {', '.join(ROLES)}")
    
    # Create user (pre-approved)
    new_user = {
        "id": str(uuid.uuid4()),
        "email": user_data.email,
        "nom": user_data.nom,
        "prenom": user_data.prenom,
        "password_hash": pwd_context.hash(user_data.password),
        "role": user_data.role,
        "is_active": True,
        "is_approved": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.users.insert_one(new_user)
    
    # Send welcome email with credentials
    await send_welcome_email(
        user_data.email, 
        f"{user_data.prenom} {user_data.nom}", 
        user_data.password
    )
    
    # Return user without password
    del new_user["password_hash"]
    return {"message": "Utilisateur créé avec succès", "user": new_user}

@api_router.put("/users/{user_id}/role")
async def update_user_role(user_id: str, role: str, admin: dict = Depends(require_admin)):
    if role not in ROLES:
        raise HTTPException(status_code=400, detail=f"Rôle invalide. Choix: {', '.join(ROLES)}")
    result = await db.users.update_one({"id": user_id}, {"$set": {"role": role}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
    return {"message": "Rôle mis à jour"}

@api_router.put("/users/{user_id}/approve")
async def approve_user(user_id: str, admin: dict = Depends(require_admin)):
    """Approve a pending user"""
    # Get user info before update for email
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    
    result = await db.users.update_one(
        {"id": user_id}, 
        {"$set": {"is_approved": True, "is_active": True}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
    
    # Send approval email
    if user:
        await send_access_approved_email(
            user["email"], 
            f"{user.get('prenom', '')} {user.get('nom', '')}"
        )
    
    return {"message": "Utilisateur approuvé"}

@api_router.put("/users/{user_id}/reject")
async def reject_user(user_id: str, admin: dict = Depends(require_admin)):
    """Reject a pending user (delete them)"""
    # Get user info before delete for email
    user = await db.users.find_one({"id": user_id, "is_approved": False}, {"_id": 0})
    
    result = await db.users.delete_one({"id": user_id, "is_approved": False})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé ou déjà approuvé")
    
    # Send rejection email
    if user:
        await send_access_rejected_email(
            user["email"], 
            f"{user.get('prenom', '')} {user.get('nom', '')}"
        )
    
    return {"message": "Demande refusée"}

@api_router.put("/users/{user_id}/suspend")
async def suspend_user(user_id: str, admin: dict = Depends(require_admin)):
    """Suspend a user"""
    # Prevent admin from suspending themselves
    if user_id == admin["id"]:
        raise HTTPException(status_code=400, detail="Vous ne pouvez pas vous suspendre vous-même")
    
    result = await db.users.update_one({"id": user_id}, {"$set": {"is_active": False}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
    return {"message": "Utilisateur suspendu"}

@api_router.put("/users/{user_id}/activate")
async def activate_user(user_id: str, admin: dict = Depends(require_admin)):
    """Reactivate a suspended user"""
    result = await db.users.update_one({"id": user_id}, {"$set": {"is_active": True}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
    return {"message": "Utilisateur réactivé"}

class PasswordChange(BaseModel):
    current_password: str  # Always required for self-service
    new_password: str

class AdminPasswordChange(BaseModel):
    new_password: str

@api_router.put("/users/me/change-password")
async def change_my_password(data: PasswordChange, current_user: dict = Depends(get_current_user)):
    """Self-service password change - requires current password verification"""
    # Get user from database
    user = await db.users.find_one({"id": current_user["id"]})
    if not user:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
    
    # Verify current password
    if not pwd_context.verify(data.current_password, user["password_hash"]):
        raise HTTPException(status_code=400, detail="Mot de passe actuel incorrect")
    
    # Validate new password
    if len(data.new_password) < 6:
        raise HTTPException(status_code=400, detail="Le nouveau mot de passe doit contenir au moins 6 caractères")
    
    # Update password
    new_hash = pwd_context.hash(data.new_password)
    await db.users.update_one({"id": current_user["id"]}, {"$set": {"password_hash": new_hash, "must_change_password": False}})
    
    return {"message": "Mot de passe modifié avec succès"}

@api_router.put("/users/{user_id}/password")
async def change_user_password(user_id: str, data: AdminPasswordChange, admin: dict = Depends(require_admin)):
    """Admin-only: Change any user's password without verification"""
    # Get user
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
    
    # Update password
    new_hash = pwd_context.hash(data.new_password)
    await db.users.update_one({"id": user_id}, {"$set": {"password_hash": new_hash}})
    
    return {"message": "Mot de passe modifié avec succès"}

@api_router.get("/users/reset-requests")
async def get_reset_requests(admin: dict = Depends(require_admin)):
    """List pending password reset requests."""
    reqs = await db.password_reset_requests.find({"statut": "pending"}, {"_id": 0}).sort("created_at", -1).to_list(200)
    return reqs

@api_router.post("/users/{user_id}/send-temp-password")
async def send_temp_password(user_id: str, admin: dict = Depends(require_admin)):
    """Admin: generate a temporary password, email it to the user, and force a change at next login."""
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
    temp = generate_temp_password()
    await db.users.update_one(
        {"id": user_id},
        {"$set": {"password_hash": hash_password(temp), "must_change_password": True}}
    )
    await db.password_reset_requests.update_many(
        {"user_id": user_id, "statut": "pending"},
        {"$set": {"statut": "resolved", "resolved_at": datetime.now(timezone.utc).isoformat(), "resolved_by": admin.get("email")}}
    )
    name = f"{user.get('prenom', '')} {user.get('nom', '')}".strip() or user["email"]
    email_sent = await send_temp_password_email(user["email"], name, temp)
    return {
        "message": f"Mot de passe temporaire envoyé à {user['email']}",
        "temp_password": temp,
        "email_sent": email_sent,
    }

@api_router.delete("/users/reset-requests/{request_id}")
async def dismiss_reset_request(request_id: str, admin: dict = Depends(require_admin)):
    """Admin: dismiss (resolve) a reset request without action."""
    result = await db.password_reset_requests.update_one(
        {"id": request_id},
        {"$set": {"statut": "resolved", "resolved_at": datetime.now(timezone.utc).isoformat(), "resolved_by": admin.get("email")}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Demande non trouvée")
    return {"message": "Demande traitée"}

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, admin: dict = Depends(require_admin)):
    # Prevent admin from deleting themselves
    if user_id == admin["id"]:
        raise HTTPException(status_code=400, detail="Vous ne pouvez pas supprimer votre propre compte")
    
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
    return {"message": "Utilisateur supprimé"}

@api_router.get("/users/permissions")
async def get_user_permissions(current_user: dict = Depends(get_current_user)):
    """Return permissions for the current user based on their role"""
    role = current_user.get("role", "invite")
    
    permissions = {
        "can_create": role in ["admin", "technicien"],
        "can_modify": role in ["admin", "technicien"],
        "can_delete": role == "admin",
        "can_export": role in ["admin", "technicien"],
        "can_manage_users": role == "admin",
        "role": role
    }
    
    return permissions

# ==================== CAISSON ROUTES ====================

@api_router.post("/caisson", response_model=Caisson)
async def create_caisson(data: CaissonCreate, current_user: dict = Depends(get_current_user)):
    existing = await db.caisson.find_one({})
    if existing:
        raise HTTPException(status_code=400, detail="Un caisson existe déjà. Utilisez PUT pour modifier.")
    
    caisson = Caisson(**data.model_dump())
    doc = caisson.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    await db.caisson.insert_one(doc)
    return caisson

@api_router.get("/caisson", response_model=Optional[Caisson])
async def get_caisson(current_user: dict = Depends(get_current_user)):
    caisson = await db.caisson.find_one({}, {"_id": 0})
    return caisson

@api_router.put("/caisson/{caisson_id}", response_model=Caisson)
async def update_caisson(caisson_id: str, data: CaissonCreate, current_user: dict = Depends(get_current_user)):
    result = await db.caisson.update_one({"id": caisson_id}, {"$set": data.model_dump()})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Caisson non trouvé")
    caisson = await db.caisson.find_one({"id": caisson_id}, {"_id": 0})
    return caisson

# ==================== EQUIPMENT TYPES ROUTES ====================

DEFAULT_EQUIPMENT_TYPES = [
    {"code": "porte", "nom": "Porte", "description": "Portes du caisson"},
    {"code": "joint", "nom": "Joint", "description": "Joints d'étanchéité"},
    {"code": "soupape", "nom": "Soupape", "description": "Soupapes de sécurité"},
    {"code": "compresseur", "nom": "Compresseur", "description": "Compresseurs d'air"},
    {"code": "capteur", "nom": "Capteur", "description": "Capteurs de pression/température"},
    {"code": "systeme_securite", "nom": "Système de sécurité", "description": "Systèmes de sécurité"},
]

@api_router.get("/equipment-types", response_model=List[EquipmentType])
async def get_equipment_types(current_user: dict = Depends(get_current_user)):
    types = await db.equipment_types.find({}, {"_id": 0}).to_list(1000)
    # Si aucun type, initialiser avec les types par défaut
    if not types:
        for t in DEFAULT_EQUIPMENT_TYPES:
            eq_type = EquipmentType(**t)
            doc = eq_type.model_dump()
            doc["created_at"] = doc["created_at"].isoformat()
            await db.equipment_types.insert_one(doc)
        types = await db.equipment_types.find({}, {"_id": 0}).to_list(1000)
    return types

@api_router.post("/equipment-types", response_model=EquipmentType)
async def create_equipment_type(data: EquipmentTypeCreate, current_user: dict = Depends(get_current_user)):
    # Vérifier que le nom n'existe pas déjà
    existing = await db.equipment_types.find_one({"nom": data.nom})
    if existing:
        raise HTTPException(status_code=400, detail="Un type avec ce nom existe déjà")
    
    eq_type = EquipmentType(**data.model_dump())
    doc = eq_type.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    await db.equipment_types.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.put("/equipment-types/{type_id}", response_model=EquipmentType)
async def update_equipment_type(type_id: str, data: EquipmentTypeCreate, current_user: dict = Depends(get_current_user)):
    result = await db.equipment_types.update_one({"id": type_id}, {"$set": data.model_dump()})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Type d'équipement non trouvé")
    eq_type = await db.equipment_types.find_one({"id": type_id}, {"_id": 0})
    return eq_type

@api_router.delete("/equipment-types/{type_id}")
async def delete_equipment_type(type_id: str, current_user: dict = Depends(get_current_user)):
    # Vérifier qu'aucun équipement n'utilise ce type
    eq_type = await db.equipment_types.find_one({"id": type_id})
    if eq_type:
        count = await db.equipments.count_documents({"type": eq_type["code"]})
        if count > 0:
            raise HTTPException(status_code=400, detail=f"Impossible de supprimer: {count} équipement(s) utilisent ce type")
    
    result = await db.equipment_types.delete_one({"id": type_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Type d'équipement non trouvé")
    return {"message": "Type d'équipement supprimé"}

# ==================== EQUIPMENT ROUTES ====================

@api_router.post("/equipments", response_model=Equipment)
async def create_equipment(data: EquipmentCreate, current_user: dict = Depends(get_current_user)):
    equipment = Equipment(**data.model_dump())
    doc = equipment.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    await db.equipments.insert_one(doc)
    return equipment

@api_router.get("/equipments", response_model=List[Equipment])
async def get_equipments(
    type: Optional[str] = None,
    statut: Optional[str] = None,
    criticite: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if type:
        query["type"] = type
    if statut:
        query["statut"] = statut
    if criticite:
        query["criticite"] = criticite
    
    equipments = await db.equipments.find(query, {"_id": 0}).to_list(1000)
    return equipments

@api_router.get("/equipments/{equipment_id}", response_model=Equipment)
async def get_equipment(equipment_id: str, current_user: dict = Depends(get_current_user)):
    equipment = await db.equipments.find_one({"id": equipment_id}, {"_id": 0})
    if not equipment:
        raise HTTPException(status_code=404, detail="Équipement non trouvé")
    return equipment

@api_router.put("/equipments/{equipment_id}", response_model=Equipment)
async def update_equipment(equipment_id: str, data: EquipmentCreate, current_user: dict = Depends(get_current_user)):
    existing = await db.equipments.find_one({"id": equipment_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Équipement non trouvé")
    update_ops = {"$set": data.model_dump()}
    # Journaliser un changement de statut (qui, quand, pourquoi)
    if data.statut != existing.get("statut"):
        entry = {
            "date": datetime.now(timezone.utc).isoformat(),
            "ancien_statut": existing.get("statut"),
            "nouveau_statut": data.statut,
            "motif": data.motif_reforme if data.statut == "reforme" else None,
            "technicien_responsable": data.technicien_reforme if data.statut == "reforme" else None,
            "utilisateur": current_user.get("email") or current_user.get("nom") or "inconnu",
        }
        update_ops["$push"] = {"historique_statut": entry}
    await db.equipments.update_one({"id": equipment_id}, update_ops)
    # Si l'équipement passe en réformé : annuler ses maintenances préventives actives
    if data.statut == "reforme" and existing.get("statut") != "reforme":
        await db.work_orders.update_many(
            {"equipment_id": equipment_id, "type_maintenance": "preventive", "statut": {"$in": ["planifiee", "en_cours"]}},
            {"$set": {"statut": "annulee"}}
        )
    equipment = await db.equipments.find_one({"id": equipment_id}, {"_id": 0})
    return equipment

@api_router.delete("/equipments/{equipment_id}")
async def delete_equipment(equipment_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.equipments.delete_one({"id": equipment_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Équipement non trouvé")
    # Supprimer aussi les sous-équipements liés
    await db.subequipments.delete_many({"parent_equipment_id": equipment_id})
    return {"message": "Équipement supprimé"}

# Route pour mettre à jour le compteur horaire d'un compresseur
class CompteurHoraireUpdate(BaseModel):
    compteur_horaire: float
    technicien: Optional[str] = None

@api_router.put("/equipments/{equipment_id}/compteur-horaire")
async def update_compteur_horaire(
    equipment_id: str, 
    data: CompteurHoraireUpdate, 
    current_user: dict = Depends(get_current_user)
):
    equipment = await db.equipments.find_one({"id": equipment_id})
    if not equipment:
        raise HTTPException(status_code=404, detail="Équipement non trouvé")
    
    if (equipment.get("type") or "").lower() != "compresseur":
        raise HTTPException(status_code=400, detail="Le compteur horaire n'est disponible que pour les compresseurs")
    
    # Ajouter à l'historique
    historique_entry = {
        "date": datetime.now(timezone.utc).isoformat(),
        "valeur": data.compteur_horaire,
        "technicien": data.technicien or current_user.get("email"),
        "ancienne_valeur": equipment.get("compteur_horaire", 0)
    }
    
    await db.equipments.update_one(
        {"id": equipment_id},
        {
            "$set": {"compteur_horaire": data.compteur_horaire},
            "$push": {"historique_compteur": historique_entry}
        }
    )
    
    # Vérifier s'il y a des maintenances préventives basées sur les heures à déclencher
    maintenances = await db.work_orders.find({
        "equipment_id": equipment_id,
        "periodicite_heures": {"$ne": None},
        "statut": {"$in": ["planifiee", "terminee"]}
    }).to_list(100)
    
    alerts = []
    for wo in maintenances:
        compteur_declenchement = wo.get("compteur_declenchement", 0)
        if data.compteur_horaire >= compteur_declenchement and wo.get("statut") == "planifiee":
            alerts.append({
                "work_order_id": wo["id"],
                "titre": wo["titre"],
                "message": f"Maintenance à effectuer: compteur {data.compteur_horaire}h >= seuil {compteur_declenchement}h"
            })
    
    updated = await db.equipments.find_one({"id": equipment_id}, {"_id": 0})
    return {"equipment": updated, "alerts": alerts}

# ==================== SUB-EQUIPMENT ROUTES ====================

@api_router.post("/subequipments", response_model=SubEquipment)
async def create_subequipment(data: SubEquipmentCreate, current_user: dict = Depends(get_current_user)):
    # Vérifier que l'équipement parent existe
    parent = await db.equipments.find_one({"id": data.parent_equipment_id})
    if not parent:
        raise HTTPException(status_code=404, detail="Équipement parent non trouvé")
    
    subequipment = SubEquipment(**data.model_dump())
    doc = subequipment.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    await db.subequipments.insert_one(doc)
    return subequipment

@api_router.get("/subequipments", response_model=List[SubEquipment])
async def get_subequipments(
    parent_equipment_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if parent_equipment_id:
        query["parent_equipment_id"] = parent_equipment_id
    
    subequipments = await db.subequipments.find(query, {"_id": 0}).to_list(1000)
    return subequipments

@api_router.get("/subequipments/{subequipment_id}", response_model=SubEquipment)
async def get_subequipment(subequipment_id: str, current_user: dict = Depends(get_current_user)):
    subequipment = await db.subequipments.find_one({"id": subequipment_id}, {"_id": 0})
    if not subequipment:
        raise HTTPException(status_code=404, detail="Sous-équipement non trouvé")
    return subequipment

@api_router.put("/subequipments/{subequipment_id}", response_model=SubEquipment)
async def update_subequipment(subequipment_id: str, data: SubEquipmentCreate, current_user: dict = Depends(get_current_user)):
    result = await db.subequipments.update_one({"id": subequipment_id}, {"$set": data.model_dump()})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Sous-équipement non trouvé")
    subequipment = await db.subequipments.find_one({"id": subequipment_id}, {"_id": 0})
    return subequipment

@api_router.delete("/subequipments/{subequipment_id}")
async def delete_subequipment(subequipment_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.subequipments.delete_one({"id": subequipment_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Sous-équipement non trouvé")
    return {"message": "Sous-équipement supprimé"}

# Sub-equipment file uploads
@api_router.post("/subequipments/{subequipment_id}/photos")
async def upload_subequipment_photo(
    subequipment_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    subequipment = await db.subequipments.find_one({"id": subequipment_id})
    if not subequipment:
        raise HTTPException(status_code=404, detail="Sous-équipement non trouvé")
    
    ext = Path(file.filename).suffix.lower()
    if ext not in {".jpg", ".jpeg", ".png", ".gif", ".webp"}:
        raise HTTPException(status_code=400, detail="Format non supporté")
    
    unique_filename = f"{uuid.uuid4()}{ext}"
    file_path = UPLOADS_DIR / "subequipments" / unique_filename
    
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    photo_url = f"/api/uploads/subequipments/{unique_filename}"
    await db.subequipments.update_one(
        {"id": subequipment_id},
        {"$push": {"photos": photo_url}}
    )
    return {"filename": file.filename, "url": photo_url}

@api_router.post("/subequipments/{subequipment_id}/documents")
async def upload_subequipment_document(
    subequipment_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    subequipment = await db.subequipments.find_one({"id": subequipment_id})
    if not subequipment:
        raise HTTPException(status_code=404, detail="Sous-équipement non trouvé")
    
    ext = Path(file.filename).suffix.lower()
    if ext != ".pdf":
        raise HTTPException(status_code=400, detail="Seuls les fichiers PDF sont acceptés")
    
    unique_filename = f"{uuid.uuid4()}{ext}"
    file_path = UPLOADS_DIR / "subequipments" / unique_filename
    
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    doc_url = f"/api/uploads/subequipments/{unique_filename}"
    doc_info = {
        "filename": file.filename,
        "url": doc_url,
        "uploaded_at": datetime.now(timezone.utc).isoformat()
    }
    await db.subequipments.update_one(
        {"id": subequipment_id},
        {"$push": {"documents": doc_info}}
    )
    return doc_info

@api_router.delete("/subequipments/{subequipment_id}/photos")
async def delete_subequipment_photo(
    subequipment_id: str,
    photo_url: str,
    current_user: dict = Depends(get_current_user)
):
    await db.subequipments.update_one(
        {"id": subequipment_id},
        {"$pull": {"photos": photo_url}}
    )
    filename = photo_url.split("/")[-1]
    file_path = UPLOADS_DIR / "subequipments" / filename
    if file_path.exists():
        file_path.unlink()
    return {"message": "Photo supprimée"}

@api_router.delete("/subequipments/{subequipment_id}/documents")
async def delete_subequipment_document(
    subequipment_id: str,
    doc_url: str,
    current_user: dict = Depends(get_current_user)
):
    await db.subequipments.update_one(
        {"id": subequipment_id},
        {"$pull": {"documents": {"url": doc_url}}}
    )
    filename = doc_url.split("/")[-1]
    file_path = UPLOADS_DIR / "subequipments" / filename
    if file_path.exists():
        file_path.unlink()
    return {"message": "Document supprimé"}

# ==================== WORK ORDER ROUTES ====================

@api_router.post("/work-orders", response_model=WorkOrder)
async def create_work_order(data: WorkOrderCreate, current_user: dict = Depends(get_current_user)):
    if data.type_maintenance == "preventive" and data.equipment_id:
        eq = await db.equipments.find_one({"id": data.equipment_id}, {"_id": 0, "statut": 1})
        if eq and eq.get("statut") == "reforme":
            raise HTTPException(status_code=400, detail="Impossible de planifier une maintenance préventive sur un équipement réformé.")
    work_order = WorkOrder(**data.model_dump())
    doc = work_order.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    await db.work_orders.insert_one(doc)
    return work_order

@api_router.get("/work-orders", response_model=List[WorkOrder])
async def get_work_orders(
    statut: Optional[str] = None,
    type_maintenance: Optional[str] = None,
    priorite: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if statut:
        query["statut"] = statut
    if type_maintenance:
        query["type_maintenance"] = type_maintenance
    if priorite:
        query["priorite"] = priorite
    
    work_orders = await db.work_orders.find(query, {"_id": 0}).to_list(1000)
    return work_orders

async def _build_maintenance_history(entity_id: str):
    """Agrège l'historique et les maintenances futures pour un équipement/sous-équipement."""
    today = datetime.now(timezone.utc).date()

    def parse(d):
        try:
            return datetime.strptime(d, "%Y-%m-%d").date()
        except (ValueError, TypeError):
            return None

    historique = []
    futures = []

    # Équipement réformé : on conserve l'historique mais plus de maintenances futures
    eq = await db.equipments.find_one({"id": entity_id}, {"_id": 0, "statut": 1})
    is_reformed = bool(eq and eq.get("statut") == "reforme")

    # Interventions réalisées (rattachées à l'équipement OU au sous-équipement)
    interventions = await db.interventions.find(
        {"$or": [{"equipment_id": entity_id}, {"sous_equipement_id": entity_id}]}, {"_id": 0}
    ).to_list(5000)
    for it in interventions:
        historique.append({
            "source": "intervention",
            "type": "Intervention " + (it.get("type_intervention") or ""),
            "titre": it.get("titre") or it.get("actions_realisees") or "Intervention",
            "date": it.get("date_intervention"),
            "statut": "terminee",
            "acteur": it.get("technicien"),
            "observations": it.get("observations"),
        })

    # Ordres de travail (maintenance préventive/corrective)
    work_orders = await db.work_orders.find({"equipment_id": entity_id}, {"_id": 0}).to_list(2000)
    for wo in work_orders:
        entry = {
            "source": "work_order",
            "type": "Maintenance " + (wo.get("type_maintenance") or "préventive"),
            "titre": wo.get("titre"),
            "statut": wo.get("statut"),
            "acteur": wo.get("technicien_assigne"),
            "observations": wo.get("description"),
            "periodicite_jours": wo.get("periodicite_jours"),
        }
        if wo.get("statut") == "terminee":
            entry["date"] = wo.get("date_realisation") or wo.get("date_planifiee")
            historique.append(entry)
        elif not is_reformed:
            entry["date"] = wo.get("date_planifiee")
            d = parse(entry["date"])
            entry["is_overdue"] = bool(d and d < today)
            futures.append(entry)

    # Contrôles réglementaires (inspections)
    inspections = await db.inspections.find({"equipment_id": entity_id}, {"_id": 0}).to_list(2000)
    for insp in inspections:
        # Dernier contrôle réalisé -> historique
        if insp.get("date_realisation"):
            historique.append({
                "source": "inspection",
                "type": "Contrôle réglementaire",
                "titre": insp.get("titre"),
                "date": insp.get("date_realisation"),
                "statut": "terminee",
                "acteur": insp.get("organisme_certificateur"),
                "observations": insp.get("resultat"),
            })
        # Prochaine échéance -> futures (sauf si équipement réformé)
        if insp.get("date_validite") and not is_reformed:
            d = parse(insp.get("date_validite"))
            futures.append({
                "source": "inspection",
                "type": "Contrôle réglementaire",
                "titre": insp.get("titre"),
                "date": insp.get("date_validite"),
                "statut": "planifiee",
                "acteur": insp.get("organisme_certificateur"),
                "periodicite": insp.get("periodicite"),
                "is_overdue": bool(d and d < today),
            })

    historique.sort(key=lambda x: x.get("date") or "", reverse=True)
    futures.sort(key=lambda x: x.get("date") or "")
    return {"historique": historique, "futures": futures}


@api_router.get("/equipments/{equipment_id}/history")
async def get_equipment_history(equipment_id: str, current_user: dict = Depends(get_current_user)):
    return await _build_maintenance_history(equipment_id)


@api_router.get("/subequipments/{subequipment_id}/history")
async def get_subequipment_history(subequipment_id: str, current_user: dict = Depends(get_current_user)):
    return await _build_maintenance_history(subequipment_id)



@api_router.get("/work-orders/{work_order_id}", response_model=WorkOrder)
async def get_work_order(work_order_id: str, current_user: dict = Depends(get_current_user)):
    work_order = await db.work_orders.find_one({"id": work_order_id}, {"_id": 0})
    if not work_order:
        raise HTTPException(status_code=404, detail="Ordre de travail non trouvé")
    return work_order

@api_router.put("/work-orders/{work_order_id}", response_model=WorkOrder)
async def update_work_order(work_order_id: str, data: WorkOrderCreate, current_user: dict = Depends(get_current_user)):
    result = await db.work_orders.update_one({"id": work_order_id}, {"$set": data.model_dump()})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Ordre de travail non trouvé")
    work_order = await db.work_orders.find_one({"id": work_order_id}, {"_id": 0})
    return work_order

@api_router.delete("/work-orders/{work_order_id}")
async def delete_work_order(work_order_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.work_orders.delete_one({"id": work_order_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Ordre de travail non trouvé")
    return {"message": "Ordre de travail supprimé"}

# Work orders file uploads
@api_router.post("/work-orders/{work_order_id}/photos")
async def upload_work_order_photo(
    work_order_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    work_order = await db.work_orders.find_one({"id": work_order_id})
    if not work_order:
        raise HTTPException(status_code=404, detail="Maintenance non trouvée")
    
    ext = Path(file.filename).suffix.lower()
    if ext not in {".jpg", ".jpeg", ".png", ".gif", ".webp"}:
        raise HTTPException(status_code=400, detail="Format non supporté")
    
    unique_filename = f"{uuid.uuid4()}{ext}"
    file_path = UPLOADS_DIR / "workorders" / unique_filename
    
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    photo_url = f"/api/uploads/workorders/{unique_filename}"
    await db.work_orders.update_one(
        {"id": work_order_id},
        {"$push": {"photos": photo_url}}
    )
    return {"filename": file.filename, "url": photo_url}

@api_router.post("/work-orders/{work_order_id}/documents")
async def upload_work_order_document(
    work_order_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    work_order = await db.work_orders.find_one({"id": work_order_id})
    if not work_order:
        raise HTTPException(status_code=404, detail="Maintenance non trouvée")
    
    ext = Path(file.filename).suffix.lower()
    if ext != ".pdf":
        raise HTTPException(status_code=400, detail="Seuls les fichiers PDF sont acceptés")
    
    unique_filename = f"{uuid.uuid4()}{ext}"
    file_path = UPLOADS_DIR / "workorders" / unique_filename
    
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    doc_url = f"/api/uploads/workorders/{unique_filename}"
    doc_info = {
        "filename": file.filename,
        "url": doc_url,
        "uploaded_at": datetime.now(timezone.utc).isoformat()
    }
    await db.work_orders.update_one(
        {"id": work_order_id},
        {"$push": {"documents": doc_info}}
    )
    return doc_info

@api_router.delete("/work-orders/{work_order_id}/photos")
async def delete_work_order_photo(
    work_order_id: str,
    photo_url: str,
    current_user: dict = Depends(get_current_user)
):
    await db.work_orders.update_one(
        {"id": work_order_id},
        {"$pull": {"photos": photo_url}}
    )
    filename = photo_url.split("/")[-1]
    file_path = UPLOADS_DIR / "workorders" / filename
    if file_path.exists():
        file_path.unlink()
    return {"message": "Photo supprimée"}

@api_router.delete("/work-orders/{work_order_id}/documents")
async def delete_work_order_document(
    work_order_id: str,
    doc_url: str,
    current_user: dict = Depends(get_current_user)
):
    await db.work_orders.update_one(
        {"id": work_order_id},
        {"$pull": {"documents": {"url": doc_url}}}
    )
    filename = doc_url.split("/")[-1]
    file_path = UPLOADS_DIR / "workorders" / filename
    if file_path.exists():
        file_path.unlink()
    return {"message": "Document supprimé"}

# ==================== INTERVENTION ROUTES ====================

@api_router.post("/interventions", response_model=Intervention)
async def create_intervention(data: InterventionCreate, current_user: dict = Depends(get_current_user)):
    # Récupérer l'équipement concerné (depuis work_order ou directement)
    equipment_id = data.equipment_id
    if not equipment_id:
        if data.work_order_id:
            wo = await db.work_orders.find_one({"id": data.work_order_id})
            if wo:
                equipment_id = wo.get("equipment_id")
        elif data.maintenance_preventive_id:
            wo = await db.work_orders.find_one({"id": data.maintenance_preventive_id})
            if wo:
                equipment_id = wo.get("equipment_id")
    
    # Mettre à jour le compteur horaire si fourni et si c'est un compresseur
    if data.compteur_horaire is not None and equipment_id:
        equipment = await db.equipments.find_one({"id": equipment_id})
        if equipment and (equipment.get("type") or "").lower() == "compresseur":
            # Mettre à jour le compteur horaire de l'équipement
            historique_entry = {
                "date": datetime.now(timezone.utc).isoformat(),
                "valeur": data.compteur_horaire,
                "technicien": data.technicien,
                "ancienne_valeur": equipment.get("compteur_horaire", 0),
                "intervention": True
            }
            await db.equipments.update_one(
                {"id": equipment_id},
                {
                    "$set": {"compteur_horaire": data.compteur_horaire},
                    "$push": {"historique_compteur": historique_entry}
                }
            )
    
    # Décrémentation du stock des pièces utilisées
    pieces_details = []
    for piece in data.pieces_utilisees:
        spare_part = await db.spare_parts.find_one({"id": piece.get("spare_part_id")})
        if spare_part:
            quantite = piece.get("quantite", 0)
            new_qty = spare_part["quantite_stock"] - quantite
            await db.spare_parts.update_one(
                {"id": piece.get("spare_part_id")},
                {"$set": {"quantite_stock": max(0, new_qty)}}
            )
            pieces_details.append({
                "spare_part_id": piece.get("spare_part_id"),
                "nom": spare_part["nom"],
                "quantite": quantite,
                "stock_avant": spare_part["quantite_stock"],
                "stock_apres": max(0, new_qty)
            })
    
    # Créer l'intervention avec l'equipment_id
    intervention_data = data.model_dump()
    intervention_data["equipment_id"] = equipment_id
    intervention = Intervention(**intervention_data)
    doc = intervention.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    await db.interventions.insert_one(doc)
    
    # Si maintenance curative (ordre de travail)
    if data.type_intervention == "curative" and data.work_order_id:
        await db.work_orders.update_one(
            {"id": data.work_order_id},
            {"$set": {"statut": "terminee"}}
        )
    
    # Si maintenance préventive, mettre à jour le work order ET recalculer la prochaine échéance
    if data.type_intervention == "preventive" and data.maintenance_preventive_id:
        # Récupérer le work order préventif
        work_order = await db.work_orders.find_one({"id": data.maintenance_preventive_id})
        if work_order:
            # Marquer comme terminée
            await db.work_orders.update_one(
                {"id": data.maintenance_preventive_id},
                {"$set": {"statut": "terminee"}}
            )
            
            # Si périodicité définie ET équipement non réformé, créer la prochaine maintenance
            eq_wo = await db.equipments.find_one({"id": work_order.get("equipment_id")}) if work_order.get("equipment_id") else None
            is_reformed = bool(eq_wo and eq_wo.get("statut") == "reforme")
            if not is_reformed and (work_order.get("periodicite_jours") or work_order.get("periodicite_heures")):
                from datetime import timedelta
                
                # Calculer la prochaine date
                if work_order.get("periodicite_jours"):
                    next_date = datetime.strptime(data.date_intervention, "%Y-%m-%d") + timedelta(days=work_order["periodicite_jours"])
                    next_date_str = next_date.strftime("%Y-%m-%d")
                else:
                    next_date_str = data.date_intervention  # Pour les heures, on garde la même date
                
                # Calculer le prochain compteur de déclenchement si basé sur les heures
                next_compteur = None
                if work_order.get("periodicite_heures") and eq_wo:
                    current_compteur = eq_wo.get("compteur_horaire", 0)
                    next_compteur = current_compteur + work_order["periodicite_heures"]
                
                # Créer le nouveau work order
                new_wo = WorkOrder(
                    titre=work_order["titre"],
                    description=work_order["description"],
                    type_maintenance="preventive",
                    priorite=work_order.get("priorite", "normale"),
                    statut="planifiee",
                    caisson_id=work_order.get("caisson_id"),
                    equipment_id=work_order.get("equipment_id"),
                    date_planifiee=next_date_str,
                    periodicite_jours=work_order.get("periodicite_jours"),
                    periodicite_heures=work_order.get("periodicite_heures"),
                    compteur_declenchement=next_compteur,
                    technicien_assigne=work_order.get("technicien_assigne")
                )
                new_doc = new_wo.model_dump()
                new_doc["created_at"] = new_doc["created_at"].isoformat()
                await db.work_orders.insert_one(new_doc)
    
    return intervention

@api_router.get("/interventions", response_model=List[Intervention])
async def get_interventions(
    work_order_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if work_order_id:
        query["work_order_id"] = work_order_id
    
    interventions = await db.interventions.find(query, {"_id": 0}).sort("date_intervention", -1).to_list(10000)
    return interventions

@api_router.get("/interventions/{intervention_id}", response_model=Intervention)
async def get_intervention(intervention_id: str, current_user: dict = Depends(get_current_user)):
    intervention = await db.interventions.find_one({"id": intervention_id}, {"_id": 0})
    if not intervention:
        raise HTTPException(status_code=404, detail="Intervention non trouvée")
    return intervention

@api_router.put("/interventions/{intervention_id}", response_model=Intervention)
async def update_intervention(intervention_id: str, data: InterventionCreate, current_user: dict = Depends(require_admin)):
    """Rectification d'une intervention (admin uniquement). Réajuste le stock des pièces."""
    existing = await db.interventions.find_one({"id": intervention_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Intervention non trouvée")

    # Réajustement du stock : re-créditer les anciennes pièces, redéduire les nouvelles
    for old in existing.get("pieces_utilisees", []):
        if old.get("spare_part_id"):
            await db.spare_parts.update_one(
                {"id": old["spare_part_id"]},
                {"$inc": {"quantite_stock": int(old.get("quantite", 0))}}
            )
    for piece in data.pieces_utilisees:
        sp = await db.spare_parts.find_one({"id": piece.get("spare_part_id")})
        if sp:
            new_qty = max(0, sp["quantite_stock"] - int(piece.get("quantite", 0)))
            await db.spare_parts.update_one(
                {"id": piece.get("spare_part_id")},
                {"$set": {"quantite_stock": new_qty}}
            )

    update = data.model_dump()
    # Conserver les champs non éditables via ce formulaire
    update["documents"] = existing.get("documents", [])
    update["equipment_id"] = data.equipment_id or existing.get("equipment_id")
    update["pieces_utilisees"] = [
        {"spare_part_id": p.get("spare_part_id"), "quantite": int(p.get("quantite", 0))}
        for p in data.pieces_utilisees
    ]
    await db.interventions.update_one({"id": intervention_id}, {"$set": update})
    updated = await db.interventions.find_one({"id": intervention_id}, {"_id": 0})
    return updated

@api_router.post("/interventions/{intervention_id}/documents")
async def upload_intervention_document(
    intervention_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Ajoute un PV/document PDF à une intervention (même après enregistrement)."""
    intervention = await db.interventions.find_one({"id": intervention_id})
    if not intervention:
        raise HTTPException(status_code=404, detail="Intervention non trouvée")

    ext = Path(file.filename).suffix.lower()
    if ext != ".pdf":
        raise HTTPException(status_code=400, detail="Seuls les fichiers PDF sont acceptés")

    unique_filename = f"{uuid.uuid4()}{ext}"
    file_path = UPLOADS_DIR / "interventions" / unique_filename
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    doc_url = f"/api/uploads/interventions/{unique_filename}"
    doc_info = {
        "filename": file.filename,
        "url": doc_url,
        "uploaded_at": datetime.now(timezone.utc).isoformat()
    }
    await db.interventions.update_one(
        {"id": intervention_id},
        {"$push": {"documents": doc_info}}
    )
    return doc_info

@api_router.delete("/interventions/{intervention_id}/documents")
async def delete_intervention_document(
    intervention_id: str,
    doc_url: str,
    current_user: dict = Depends(get_current_user)
):
    await db.interventions.update_one(
        {"id": intervention_id},
        {"$pull": {"documents": {"url": doc_url}}}
    )
    filename = doc_url.split("/")[-1]
    file_path = UPLOADS_DIR / "interventions" / filename
    if file_path.exists():
        file_path.unlink()
    return {"message": "Document supprimé"}

# ==================== INSPECTION ROUTES ====================

def calculate_next_date(date_realisation: str, periodicite: str) -> str:
    """Calculate next inspection date based on periodicity"""
    if not date_realisation:
        # Si pas de date de réalisation, partir d'aujourd'hui
        base_date = datetime.now(timezone.utc).date()
    else:
        base_date = datetime.strptime(date_realisation, "%Y-%m-%d").date()
    
    days = PERIODICITES.get(periodicite, 365)
    next_date = base_date + timedelta(days=days)
    return next_date.strftime("%Y-%m-%d")

@api_router.post("/inspections", response_model=Inspection)
async def create_inspection(data: InspectionCreate, current_user: dict = Depends(get_current_user)):
    data_dict = data.model_dump()
    # Calculer automatiquement la date de validité
    data_dict["date_validite"] = calculate_next_date(data_dict.get("date_realisation"), data_dict.get("periodicite", "annuel"))
    
    inspection = Inspection(**data_dict)
    doc = inspection.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    await db.inspections.insert_one(doc)
    return inspection

@api_router.get("/inspections", response_model=List[Inspection])
async def get_inspections(current_user: dict = Depends(get_current_user)):
    inspections = await db.inspections.find({}, {"_id": 0}).to_list(1000)
    return inspections

@api_router.get("/inspections/{inspection_id}", response_model=Inspection)
async def get_inspection(inspection_id: str, current_user: dict = Depends(get_current_user)):
    inspection = await db.inspections.find_one({"id": inspection_id}, {"_id": 0})
    if not inspection:
        raise HTTPException(status_code=404, detail="Contrôle non trouvé")
    return inspection

@api_router.put("/inspections/{inspection_id}", response_model=Inspection)
async def update_inspection(inspection_id: str, data: InspectionCreate, current_user: dict = Depends(get_current_user)):
    data_dict = data.model_dump()
    data_dict.pop("historique_controles", None)  # ne pas écraser l'historique via une simple modification
    # Recalculer la date de validité si date_realisation ou periodicite change
    data_dict["date_validite"] = calculate_next_date(data_dict.get("date_realisation"), data_dict.get("periodicite", "annuel"))
    
    result = await db.inspections.update_one({"id": inspection_id}, {"$set": data_dict})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Contrôle non trouvé")
    inspection = await db.inspections.find_one({"id": inspection_id}, {"_id": 0})
    return inspection

class RenewInspectionRequest(BaseModel):
    date_realisation: str
    resultat: Optional[str] = None
    organisme_certificateur: Optional[str] = None
    observations: Optional[str] = None

@api_router.post("/inspections/{inspection_id}/renew", response_model=Inspection)
async def renew_inspection(inspection_id: str, data: RenewInspectionRequest, current_user: dict = Depends(get_current_user)):
    """Enregistre le renouvellement d'un contrôle : archive la réalisation courante dans l'historique
    (traçabilité) puis met à jour la date de réalisation et recalcule l'échéance."""
    insp = await db.inspections.find_one({"id": inspection_id}, {"_id": 0})
    if not insp:
        raise HTTPException(status_code=404, detail="Contrôle non trouvé")

    ops = {"$set": {}}
    # Archiver la réalisation courante si elle existe
    if insp.get("date_realisation"):
        ops["$push"] = {"historique_controles": {
            "date_realisation": insp.get("date_realisation"),
            "date_validite": insp.get("date_validite"),
            "resultat": insp.get("resultat"),
            "organisme_certificateur": insp.get("organisme_certificateur"),
            "observations": insp.get("observations"),
            "archived_at": datetime.now(timezone.utc).isoformat(),
            "archived_by": current_user.get("email"),
        }}

    ops["$set"]["date_realisation"] = data.date_realisation
    ops["$set"]["date_validite"] = calculate_next_date(data.date_realisation, insp.get("periodicite", "annuel"))
    if data.resultat is not None:
        ops["$set"]["resultat"] = data.resultat
    if data.organisme_certificateur is not None:
        ops["$set"]["organisme_certificateur"] = data.organisme_certificateur
    if data.observations is not None:
        ops["$set"]["observations"] = data.observations

    await db.inspections.update_one({"id": inspection_id}, ops)
    updated = await db.inspections.find_one({"id": inspection_id}, {"_id": 0})
    return updated

@api_router.delete("/inspections/{inspection_id}")
async def delete_inspection(inspection_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.inspections.delete_one({"id": inspection_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Contrôle non trouvé")
    return {"message": "Contrôle supprimé"}

# ==================== SPARE PARTS ROUTES ====================

@api_router.post("/spare-parts", response_model=SparePart)
async def create_spare_part(data: SparePartCreate, current_user: dict = Depends(get_current_user)):
    spare_part = SparePart(**data.model_dump())
    doc = spare_part.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    await db.spare_parts.insert_one(doc)
    return spare_part

@api_router.get("/spare-parts", response_model=List[SparePart])
async def get_spare_parts(
    equipment_type: Optional[str] = None,
    low_stock: Optional[bool] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if equipment_type:
        query["equipment_type"] = equipment_type
    
    spare_parts = await db.spare_parts.find(query, {"_id": 0}).to_list(1000)
    
    if low_stock:
        spare_parts = [p for p in spare_parts if p["quantite_stock"] <= p["seuil_minimum"]]
    
    return spare_parts

@api_router.get("/spare-parts/{spare_part_id}", response_model=SparePart)
async def get_spare_part(spare_part_id: str, current_user: dict = Depends(get_current_user)):
    spare_part = await db.spare_parts.find_one({"id": spare_part_id}, {"_id": 0})
    if not spare_part:
        raise HTTPException(status_code=404, detail="Pièce non trouvée")
    return spare_part

@api_router.put("/spare-parts/{spare_part_id}", response_model=SparePart)
async def update_spare_part(spare_part_id: str, data: SparePartUpdate, current_user: dict = Depends(get_current_user)):
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Aucune donnée à mettre à jour")
    
    result = await db.spare_parts.update_one({"id": spare_part_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Pièce non trouvée")
    spare_part = await db.spare_parts.find_one({"id": spare_part_id}, {"_id": 0})
    return spare_part

@api_router.delete("/spare-parts/{spare_part_id}")
async def delete_spare_part(spare_part_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.spare_parts.delete_one({"id": spare_part_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Pièce non trouvée")
    return {"message": "Pièce supprimée"}

# Spare parts file uploads
@api_router.post("/spare-parts/{spare_part_id}/photos")
async def upload_spare_part_photo(
    spare_part_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    spare_part = await db.spare_parts.find_one({"id": spare_part_id})
    if not spare_part:
        raise HTTPException(status_code=404, detail="Pièce non trouvée")
    
    ext = Path(file.filename).suffix.lower()
    if ext not in {".jpg", ".jpeg", ".png", ".gif", ".webp"}:
        raise HTTPException(status_code=400, detail="Format non supporté")
    
    unique_filename = f"{uuid.uuid4()}{ext}"
    file_path = UPLOADS_DIR / "spareparts" / unique_filename
    
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    photo_url = f"/api/uploads/spareparts/{unique_filename}"
    await db.spare_parts.update_one(
        {"id": spare_part_id},
        {"$push": {"photos": photo_url}}
    )
    return {"filename": file.filename, "url": photo_url}

@api_router.post("/spare-parts/{spare_part_id}/documents")
async def upload_spare_part_document(
    spare_part_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    spare_part = await db.spare_parts.find_one({"id": spare_part_id})
    if not spare_part:
        raise HTTPException(status_code=404, detail="Pièce non trouvée")
    
    ext = Path(file.filename).suffix.lower()
    if ext != ".pdf":
        raise HTTPException(status_code=400, detail="Seuls les fichiers PDF sont acceptés")
    
    unique_filename = f"{uuid.uuid4()}{ext}"
    file_path = UPLOADS_DIR / "spareparts" / unique_filename
    
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    doc_url = f"/api/uploads/spareparts/{unique_filename}"
    doc_info = {
        "filename": file.filename,
        "url": doc_url,
        "uploaded_at": datetime.now(timezone.utc).isoformat()
    }
    await db.spare_parts.update_one(
        {"id": spare_part_id},
        {"$push": {"documents": doc_info}}
    )
    return doc_info

@api_router.delete("/spare-parts/{spare_part_id}/photos")
async def delete_spare_part_photo(
    spare_part_id: str,
    photo_url: str,
    current_user: dict = Depends(get_current_user)
):
    await db.spare_parts.update_one(
        {"id": spare_part_id},
        {"$pull": {"photos": photo_url}}
    )
    filename = photo_url.split("/")[-1]
    file_path = UPLOADS_DIR / "spareparts" / filename
    if file_path.exists():
        file_path.unlink()
    return {"message": "Photo supprimée"}

@api_router.delete("/spare-parts/{spare_part_id}/documents")
async def delete_spare_part_document(
    spare_part_id: str,
    doc_url: str,
    current_user: dict = Depends(get_current_user)
):
    await db.spare_parts.update_one(
        {"id": spare_part_id},
        {"$pull": {"documents": {"url": doc_url}}}
    )
    filename = doc_url.split("/")[-1]
    file_path = UPLOADS_DIR / "spareparts" / filename
    if file_path.exists():
        file_path.unlink()
    return {"message": "Document supprimé"}

# ==================== DASHBOARD / ALERTS ROUTES ====================

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(current_user: dict = Depends(get_current_user)):
    # Count equipments by status
    equipments = await db.equipments.find({}, {"_id": 0}).to_list(1000)
    equipment_stats = {
        "total": len(equipments),
        "en_service": len([e for e in equipments if e.get("statut") == "en_service"]),
        "maintenance": len([e for e in equipments if e.get("statut") == "maintenance"]),
        "hors_service": len([e for e in equipments if e.get("statut") == "hors_service"])
    }
    
    # Work orders stats
    work_orders = await db.work_orders.find({}, {"_id": 0}).to_list(1000)
    work_order_stats = {
        "total": len(work_orders),
        "planifiee": len([w for w in work_orders if w.get("statut") == "planifiee"]),
        "en_cours": len([w for w in work_orders if w.get("statut") == "en_cours"]),
        "terminee": len([w for w in work_orders if w.get("statut") == "terminee"])
    }
    
    # Spare parts with low stock
    spare_parts = await db.spare_parts.find({}, {"_id": 0}).to_list(1000)
    low_stock_parts = [p for p in spare_parts if p["quantite_stock"] <= p["seuil_minimum"]]
    
    # Compresseurs avec compteur horaire
    compresseurs = [e for e in equipments if (e.get("type") or "").lower() == "compresseur"]
    compresseurs_stats = []
    for comp in compresseurs:
        compresseurs_stats.append({
            "id": comp.get("id"),
            "reference": comp.get("reference"),
            "numero_serie": comp.get("numero_serie"),
            "compteur_horaire": comp.get("compteur_horaire", 0),
            "statut": comp.get("statut")
        })
    
    return {
        "equipment_stats": equipment_stats,
        "work_order_stats": work_order_stats,
        "low_stock_count": len(low_stock_parts),
        "total_spare_parts": len(spare_parts),
        "compresseurs": compresseurs_stats
    }

async def _reformed_equipment_ids():
    """Retourne l'ensemble des ids d'équipements réformés (exclus des maintenances futures / en retard)."""
    docs = await db.equipments.find({"statut": "reforme"}, {"id": 1, "_id": 0}).to_list(2000)
    return {d["id"] for d in docs}


@api_router.get("/dashboard/alerts")
async def get_alerts(current_user: dict = Depends(get_current_user)):
    alerts = []
    today = datetime.now(timezone.utc).date()
    reformed = await _reformed_equipment_ids()
    
    # Low stock alerts
    spare_parts = await db.spare_parts.find({}, {"_id": 0}).to_list(1000)
    for part in spare_parts:
        if part["quantite_stock"] <= part["seuil_minimum"]:
            alerts.append({
                "type": "stock_bas",
                "severity": "warning",
                "title": f"Stock bas: {part['nom']}",
                "description": f"Quantité: {part['quantite_stock']} / Seuil: {part['seuil_minimum']}",
                "item_id": part["id"],
                "item_type": "spare_part"
            })
    
    # Inspection expiration alerts (30 days before)
    inspections = await db.inspections.find({}, {"_id": 0}).to_list(2000)
    for inspection in inspections:
        if not inspection.get("date_validite"):
            continue
        if inspection.get("equipment_id") in reformed:
            continue
        try:
            expiry_date = datetime.strptime(inspection["date_validite"], "%Y-%m-%d").date()
            days_until_expiry = (expiry_date - today).days
            
            if days_until_expiry < 0:
                alerts.append({
                    "type": "controle_expire",
                    "severity": "critical",
                    "title": f"Contrôle expiré: {inspection['titre']}",
                    "description": f"Expiré depuis {abs(days_until_expiry)} jours",
                    "item_id": inspection["id"],
                    "item_type": "inspection"
                })
            elif days_until_expiry <= 30:
                alerts.append({
                    "type": "controle_proche",
                    "severity": "warning",
                    "title": f"Contrôle à renouveler: {inspection['titre']}",
                    "description": f"Expire dans {days_until_expiry} jours",
                    "item_id": inspection["id"],
                    "item_type": "inspection"
                })
        except (ValueError, KeyError, TypeError):
            pass
    
    # Overdue work orders
    work_orders = await db.work_orders.find({"statut": {"$in": ["planifiee", "en_cours"]}}, {"_id": 0}).to_list(1000)
    for wo in work_orders:
        if wo.get("equipment_id") in reformed:
            continue
        try:
            planned_date = datetime.strptime(wo["date_planifiee"], "%Y-%m-%d").date()
            if planned_date < today:
                days_overdue = (today - planned_date).days
                alerts.append({
                    "type": "maintenance_retard",
                    "severity": "critical" if days_overdue > 7 else "warning",
                    "title": f"Maintenance en retard: {wo['titre']}",
                    "description": f"En retard de {days_overdue} jours",
                    "item_id": wo["id"],
                    "item_type": "work_order"
                })
        except (ValueError, KeyError):
            pass
    
    # Equipment out of service
    equipments = await db.equipments.find({"statut": "hors_service"}, {"_id": 0}).to_list(1000)
    for eq in equipments:
        alerts.append({
            "type": "equipement_hs",
            "severity": "critical",
            "title": f"Équipement hors service: {eq['type']}",
            "description": f"Réf: {eq['reference']} - S/N: {eq['numero_serie']}",
            "item_id": eq["id"],
            "item_type": "equipment"
        })
    
    # Gas cylinder alerts
    gas_cylinders = await db.gas_cylinders.find({}, {"_id": 0}).to_list(1000)
    for cyl in gas_cylinders:
        # Gas expiration
        if cyl.get("date_expiration_gaz"):
            try:
                exp_date = datetime.strptime(cyl["date_expiration_gaz"], "%Y-%m-%d").date()
                days_left = (exp_date - today).days
                if days_left < 0:
                    alerts.append({
                        "type": "gaz_expire",
                        "severity": "critical",
                        "title": f"Gaz expiré: {cyl['numero_bouteille']}",
                        "description": f"{cyl['type_gaz']} - Expiré depuis {abs(days_left)} jours",
                        "item_id": cyl["id"],
                        "item_type": "gas_cylinder"
                    })
                elif days_left <= 30:
                    alerts.append({
                        "type": "gaz_expire_bientot",
                        "severity": "warning",
                        "title": f"Gaz expire bientôt: {cyl['numero_bouteille']}",
                        "description": f"{cyl['type_gaz']} - Expire dans {days_left} jours",
                        "item_id": cyl["id"],
                        "item_type": "gas_cylinder"
                    })
            except (ValueError, KeyError):
                pass
        
        # Hydraulic test expiration
        if cyl.get("date_prochaine_epreuve"):
            try:
                epr_date = datetime.strptime(cyl["date_prochaine_epreuve"], "%Y-%m-%d").date()
                days_left = (epr_date - today).days
                if days_left < 0:
                    alerts.append({
                        "type": "epreuve_expiree",
                        "severity": "critical",
                        "title": f"Épreuve expirée: {cyl['numero_bouteille']}",
                        "description": f"{cyl['type_gaz']} - Requalification requise",
                        "item_id": cyl["id"],
                        "item_type": "gas_cylinder"
                    })
                elif days_left <= 90:
                    alerts.append({
                        "type": "epreuve_bientot",
                        "severity": "warning",
                        "title": f"Épreuve dans {days_left}j: {cyl['numero_bouteille']}",
                        "description": f"{cyl['type_gaz']} - Prévoir requalification",
                        "item_id": cyl["id"],
                        "item_type": "gas_cylinder"
                    })
            except (ValueError, KeyError):
                pass
    
    # Contract expiration alerts
    contracts = await db.contracts.find({"statut": "actif"}, {"_id": 0}).to_list(1000)
    for contract in contracts:
        if contract.get("date_fin"):
            try:
                end_date = datetime.strptime(contract["date_fin"], "%Y-%m-%d").date()
                days_left = (end_date - today).days
                if days_left < 0:
                    alerts.append({
                        "type": "contrat_expire",
                        "severity": "warning",
                        "title": f"Contrat expiré: {contract['titre']}",
                        "description": f"Expiré depuis {abs(days_left)} jours",
                        "item_id": contract["id"],
                        "item_type": "contract"
                    })
                elif days_left <= 60:
                    alerts.append({
                        "type": "contrat_expire_bientot",
                        "severity": "info",
                        "title": f"Contrat expire bientôt: {contract['titre']}",
                        "description": f"Expire dans {days_left} jours",
                        "item_id": contract["id"],
                        "item_type": "contract"
                    })
            except (ValueError, KeyError):
                pass
    
    # Sort by severity
    severity_order = {"critical": 0, "warning": 1, "info": 2}
    alerts.sort(key=lambda x: severity_order.get(x["severity"], 3))
    
    return alerts

@api_router.get("/dashboard/upcoming-maintenance")
async def get_upcoming_maintenance(current_user: dict = Depends(get_current_user)):
    today = datetime.now(timezone.utc).date()
    reformed = await _reformed_equipment_ids()
    work_orders = await db.work_orders.find(
        {"statut": {"$in": ["planifiee", "en_cours"]}},
        {"_id": 0}
    ).to_list(1000)
    
    upcoming = []
    for wo in work_orders:
        if wo.get("equipment_id") in reformed:
            continue
        try:
            planned_date = datetime.strptime(wo["date_planifiee"], "%Y-%m-%d").date()
            days_diff = (planned_date - today).days
            wo["days_until"] = days_diff
            wo["is_overdue"] = days_diff < 0
            wo["origine"] = "ordre_travail"
            upcoming.append(wo)
        except (ValueError, KeyError, TypeError):
            pass

    # Inclure les contrôles réglementaires / maintenances préventives (prochaine échéance)
    inspections = await db.inspections.find({}, {"_id": 0}).to_list(3000)
    for insp in inspections:
        dv = insp.get("date_validite")
        if not dv:
            continue
        if insp.get("equipment_id") in reformed:
            continue
        try:
            planned_date = datetime.strptime(dv, "%Y-%m-%d").date()
        except (ValueError, TypeError):
            continue
        days_diff = (planned_date - today).days
        upcoming.append({
            "id": insp["id"],
            "titre": insp.get("titre"),
            "date_planifiee": dv,
            "type_maintenance": "preventive",
            "statut": "planifiee",
            "equipment_id": insp.get("equipment_id"),
            "periodicite": insp.get("periodicite"),
            "days_until": days_diff,
            "is_overdue": days_diff < 0,
            "origine": "controle_reglementaire",
        })

    # Sort by date
    upcoming.sort(key=lambda x: x.get("days_until", 999))
    
    return upcoming[:10]  # Return next 10

@api_router.get("/dashboard/calendar")
async def get_maintenance_calendar(current_user: dict = Depends(get_current_user)):
    """Retourne les maintenances planifiées sur 52 semaines pour le calendrier"""
    today = datetime.now(timezone.utc).date()
    end_date = today + timedelta(weeks=52)
    reformed = await _reformed_equipment_ids()
    
    work_orders = await db.work_orders.find(
        {"statut": {"$in": ["planifiee", "en_cours", "terminee"]}},
        {"_id": 0}
    ).to_list(1000)
    
    calendar_data = []
    
    for wo in work_orders:
        if wo.get("equipment_id") in reformed:
            continue
        try:
            planned_date = datetime.strptime(wo["date_planifiee"], "%Y-%m-%d").date()
            # Inclure les maintenances passées (4 semaines) et futures (52 semaines)
            if planned_date >= today - timedelta(weeks=4) and planned_date <= end_date:
                # Calculer le numéro de semaine
                week_number = planned_date.isocalendar()[1]
                year = planned_date.year
                
                calendar_data.append({
                    "id": wo["id"],
                    "titre": wo["titre"],
                    "type_maintenance": wo.get("type_maintenance", "preventive"),
                    "date_planifiee": wo["date_planifiee"],
                    "statut": wo["statut"],
                    "equipment_id": wo.get("equipment_id"),
                    "priorite": wo.get("priorite", "normale"),
                    "week_number": week_number,
                    "year": year,
                    "periodicite_jours": wo.get("periodicite_jours"),
                    "periodicite_heures": wo.get("periodicite_heures")
                })
        except (ValueError, KeyError, TypeError):
            pass

    # Inclure les contrôles réglementaires (prochaine échéance)
    inspections = await db.inspections.find({}, {"_id": 0}).to_list(3000)
    for insp in inspections:
        dv = insp.get("date_validite")
        if not dv:
            continue
        if insp.get("equipment_id") in reformed:
            continue
        try:
            planned_date = datetime.strptime(dv, "%Y-%m-%d").date()
        except (ValueError, TypeError):
            continue
        if today - timedelta(weeks=4) <= planned_date <= end_date:
            calendar_data.append({
                "id": insp["id"],
                "titre": insp.get("titre"),
                "type_maintenance": "preventive",
                "date_planifiee": dv,
                "statut": "planifiee",
                "equipment_id": insp.get("equipment_id"),
                "priorite": "normale",
                "week_number": planned_date.isocalendar()[1],
                "year": planned_date.year,
                "periodicite_jours": None,
                "periodicite_heures": None,
            })

    # Formations — une entrée par formation (à sa date de début), avec plage
    role = current_user.get("role")
    uid = current_user.get("id")
    formations = await db.formations.find({}, {"_id": 0}).to_list(1000)
    for f in formations:
        if role != "admin" and f.get("technicien_id") != uid:
            continue
        try:
            d0 = datetime.strptime(f["date_debut"][:10], "%Y-%m-%d").date()
        except (ValueError, TypeError, KeyError):
            continue
        if today - timedelta(weeks=4) <= d0 <= end_date:
            calendar_data.append({
                "id": f["id"],
                "titre": f"Formation : {f.get('nom')} ({f.get('technicien')})",
                "type_maintenance": "formation",
                "date_planifiee": f["date_debut"],
                "date_fin": f.get("date_fin"),
                "statut": "planifiee",
                "equipment_id": None,
                "priorite": "normale",
                "is_formation": True,
                "week_number": d0.isocalendar()[1],
                "year": d0.year,
            })

    # Trier par date
    calendar_data.sort(key=lambda x: x["date_planifiee"])
    
    return calendar_data

# ==================== PLANNING / SCHEDULING ROUTES ====================

class CompleteWorkOrderRequest(BaseModel):
    date_realisation: Optional[str] = None
    technicien: Optional[str] = None
    observations: Optional[str] = None
    compteur_horaire: Optional[float] = None

class RescheduleRequest(BaseModel):
    item_type: str  # "work_order" ou "inspection"
    item_id: str
    new_date: str  # YYYY-MM-DD


@api_router.post("/work-orders/{work_order_id}/complete")
async def complete_work_order(work_order_id: str, data: CompleteWorkOrderRequest, current_user: dict = Depends(get_current_user)):
    """Marque un ordre de maintenance comme terminé et génère automatiquement la prochaine occurrence."""
    wo = await db.work_orders.find_one({"id": work_order_id}, {"_id": 0})
    if not wo:
        raise HTTPException(status_code=404, detail="Ordre de travail non trouvé")

    date_real = data.date_realisation or datetime.now(timezone.utc).strftime("%Y-%m-%d")

    # Marquer comme terminé
    await db.work_orders.update_one(
        {"id": work_order_id},
        {"$set": {"statut": "terminee", "date_realisation": date_real}}
    )

    # Enregistrer une intervention pour tracer la réalisation
    intervention = {
        "id": str(uuid.uuid4()),
        "work_order_id": work_order_id,
        "maintenance_preventive_id": None,
        "type_intervention": "preventive",
        "date_intervention": date_real,
        "technicien": data.technicien or "Non renseigné",
        "actions_realisees": wo.get("titre", "Maintenance réalisée"),
        "observations": data.observations,
        "pieces_utilisees": [],
        "duree_minutes": None,
        "compteur_horaire": data.compteur_horaire,
        "equipment_id": wo.get("equipment_id"),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.interventions.insert_one({k: v for k, v in intervention.items()})

    # Mise à jour compteur horaire compresseur si fourni
    if data.compteur_horaire is not None and wo.get("equipment_id"):
        eq = await db.equipments.find_one({"id": wo["equipment_id"]})
        if eq and (eq.get("type") or "").lower() == "compresseur":
            await db.equipments.update_one(
                {"id": wo["equipment_id"]},
                {"$set": {"compteur_horaire": data.compteur_horaire},
                 "$push": {"historique_compteur": {
                     "date": datetime.now(timezone.utc).isoformat(),
                     "valeur": data.compteur_horaire,
                     "technicien": data.technicien or "Non renseigné"}}}
            )

    # Générer la prochaine occurrence (préventive + périodicité en jours)
    next_wo = None
    if wo.get("type_maintenance") == "preventive" and wo.get("periodicite_jours"):
        try:
            base = datetime.strptime(date_real, "%Y-%m-%d").date()
            next_date = (base + timedelta(days=int(wo["periodicite_jours"]))).strftime("%Y-%m-%d")
            next_wo = {
                "id": str(uuid.uuid4()),
                "titre": wo.get("titre"),
                "description": wo.get("description"),
                "type_maintenance": "preventive",
                "priorite": wo.get("priorite", "normale"),
                "statut": "planifiee",
                "caisson_id": wo.get("caisson_id"),
                "equipment_id": wo.get("equipment_id"),
                "date_planifiee": next_date,
                "periodicite_jours": wo.get("periodicite_jours"),
                "periodicite_heures": wo.get("periodicite_heures"),
                "compteur_declenchement": None,
                "technicien_assigne": wo.get("technicien_assigne"),
                "photos": [],
                "documents": [],
                "source": "auto_generated",
                "parent_work_order_id": work_order_id,
                "created_at": datetime.now(timezone.utc).isoformat(),
            }
            await db.work_orders.insert_one({k: v for k, v in next_wo.items()})
            next_wo.pop("_id", None)
        except (ValueError, TypeError):
            next_wo = None

    return {"completed": True, "work_order_id": work_order_id, "date_realisation": date_real,
            "next_work_order": next_wo}


@api_router.post("/planning/reschedule")
async def reschedule_maintenance(data: RescheduleRequest, current_user: dict = Depends(get_current_user)):
    """Replanifie une maintenance (glisser-déposer dans le calendrier)."""
    try:
        datetime.strptime(data.new_date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Date invalide (format attendu: YYYY-MM-DD)")

    if data.item_type == "work_order":
        result = await db.work_orders.update_one(
            {"id": data.item_id}, {"$set": {"date_planifiee": data.new_date}})
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Ordre de travail non trouvé")
    elif data.item_type == "inspection":
        result = await db.inspections.update_one(
            {"id": data.item_id}, {"$set": {"date_validite": data.new_date}})
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Contrôle non trouvé")
    else:
        raise HTTPException(status_code=400, detail="item_type invalide")

    return {"success": True, "item_id": data.item_id, "new_date": data.new_date}


@api_router.get("/planning/events")
async def get_planning_events(start: str, end: str, equipment_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """Retourne toutes les maintenances (ordres de travail + contrôles réglementaires) sur une plage de dates."""
    try:
        start_date = datetime.strptime(start, "%Y-%m-%d").date()
        end_date = datetime.strptime(end, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Dates invalides (format attendu: YYYY-MM-DD)")

    today = datetime.now(timezone.utc).date()
    events = []
    reformed = await _reformed_equipment_ids()
    wo_filter = {"equipment_id": equipment_id} if equipment_id else {}

    # Ordres de travail (maintenance préventive/corrective)
    work_orders = await db.work_orders.find(wo_filter, {"_id": 0}).to_list(3000)
    for wo in work_orders:
        if wo.get("equipment_id") in reformed:
            continue
        dp = wo.get("date_planifiee")
        if not dp:
            continue
        try:
            d = datetime.strptime(dp, "%Y-%m-%d").date()
        except (ValueError, TypeError):
            continue
        if not (start_date <= d <= end_date):
            continue
        events.append({
            "id": wo["id"],
            "item_type": "work_order",
            "origine": "preventive" if wo.get("type_maintenance") == "preventive" else "corrective",
            "titre": wo.get("titre"),
            "date": dp,
            "statut": wo.get("statut", "planifiee"),
            "equipment_id": wo.get("equipment_id"),
            "priorite": wo.get("priorite", "normale"),
            "periodicite_jours": wo.get("periodicite_jours"),
            "is_overdue": d < today and wo.get("statut") != "terminee",
        })

    # Contrôles réglementaires (inspections) via date_validite
    insp_filter = {"equipment_id": equipment_id} if equipment_id else {}
    inspections = await db.inspections.find(insp_filter, {"_id": 0}).to_list(3000)
    for insp in inspections:
        dv = insp.get("date_validite")
        if not dv:
            continue
        if insp.get("equipment_id") in reformed:
            continue
        try:
            d = datetime.strptime(dv, "%Y-%m-%d").date()
        except (ValueError, TypeError):
            continue
        if not (start_date <= d <= end_date):
            continue
        events.append({
            "id": insp["id"],
            "item_type": "inspection",
            "origine": "reglementaire",
            "titre": insp.get("titre"),
            "date": dv,
            "statut": "planifiee",
            "equipment_id": insp.get("equipment_id"),
            "priorite": "normale",
            "periodicite": insp.get("periodicite"),
            "is_overdue": d < today,
        })

    # Formations (créneaux de formation) — bande sur chaque jour couvert
    role = current_user.get("role")
    uid = current_user.get("id")
    formations = await db.formations.find({}, {"_id": 0}).to_list(1000)
    for f in formations:
        if role != "admin" and f.get("technicien_id") != uid:
            continue
        for day in _formation_event_days(f, start_date, end_date):
            ds = day.strftime("%Y-%m-%d")
            events.append({
                "id": f"formation-{f['id']}-{ds}",
                "formation_id": f["id"],
                "item_type": "formation",
                "origine": "formation",
                "titre": f"Formation : {f.get('nom')} ({f.get('technicien')})",
                "date": ds,
                "statut": "planifiee",
                "equipment_id": None,
                "priorite": "normale",
                "is_overdue": False,
            })

    events.sort(key=lambda x: x["date"])
    return events


@api_router.get("/planning/summary")
async def get_planning_summary(year: int, equipment_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """Compte des maintenances par mois pour une année (vue annuelle)."""
    start_date = datetime(year, 1, 1).date()
    end_date = datetime(year, 12, 31).date()
    today = datetime.now(timezone.utc).date()

    months = {m: {"preventive": 0, "reglementaire": 0, "overdue": 0} for m in range(1, 13)}
    ent_filter = {"equipment_id": equipment_id} if equipment_id else {}
    reformed = await _reformed_equipment_ids()

    work_orders = await db.work_orders.find(ent_filter, {"_id": 0}).to_list(3000)
    for wo in work_orders:
        if wo.get("equipment_id") in reformed:
            continue
        dp = wo.get("date_planifiee")
        if not dp:
            continue
        try:
            d = datetime.strptime(dp, "%Y-%m-%d").date()
        except (ValueError, TypeError):
            continue
        if start_date <= d <= end_date:
            months[d.month]["preventive"] += 1
            if d < today and wo.get("statut") != "terminee":
                months[d.month]["overdue"] += 1

    inspections = await db.inspections.find(ent_filter, {"_id": 0}).to_list(3000)
    for insp in inspections:
        dv = insp.get("date_validite")
        if not dv:
            continue
        if insp.get("equipment_id") in reformed:
            continue
        try:
            d = datetime.strptime(dv, "%Y-%m-%d").date()
        except (ValueError, TypeError):
            continue
        if start_date <= d <= end_date:
            months[d.month]["reglementaire"] += 1
            if d < today:
                months[d.month]["overdue"] += 1

    return {"year": year, "months": months}


# ==================== FORMATIONS ROUTES ====================

def _formation_event_days(f: dict, window_start=None, window_end=None):
    """Génère la liste des jours (date) couverts par une formation, éventuellement bornée."""
    try:
        d0 = datetime.strptime(f["date_debut"][:10], "%Y-%m-%d").date()
        d1 = datetime.strptime(f["date_fin"][:10], "%Y-%m-%d").date()
    except (ValueError, TypeError, KeyError):
        return []
    if d1 < d0:
        d0, d1 = d1, d0
    if window_start:
        d0 = max(d0, window_start)
    if window_end:
        d1 = min(d1, window_end)
    days = []
    cur = d0
    while cur <= d1:
        days.append(cur)
        cur += timedelta(days=1)
    return days


@api_router.post("/formations", response_model=Formation)
async def create_formation(data: FormationCreate, current_user: dict = Depends(require_admin)):
    formation = Formation(**data.model_dump())
    doc = formation.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    await db.formations.insert_one(doc)
    return formation


@api_router.get("/formations", response_model=List[Formation])
async def get_formations(current_user: dict = Depends(get_current_user)):
    """Admin voit tout ; les autres ne voient que leurs propres formations."""
    query = {}
    if current_user.get("role") != "admin":
        query = {"technicien_id": current_user.get("id")}
    formations = await db.formations.find(query, {"_id": 0}).sort("date_debut", -1).to_list(1000)
    return formations


@api_router.put("/formations/{formation_id}", response_model=Formation)
async def update_formation(formation_id: str, data: FormationCreate, current_user: dict = Depends(require_admin)):
    result = await db.formations.update_one({"id": formation_id}, {"$set": data.model_dump()})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Formation non trouvée")
    formation = await db.formations.find_one({"id": formation_id}, {"_id": 0})
    return formation


@api_router.delete("/formations/{formation_id}")
async def delete_formation(formation_id: str, current_user: dict = Depends(require_admin)):
    result = await db.formations.delete_one({"id": formation_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Formation non trouvée")
    return {"message": "Formation supprimée"}


@api_router.get("/search")
async def global_search(q: str, current_user: dict = Depends(get_current_user)):
    """Recherche globale: équipements, sous-équipements, maintenances, interventions, contrôles."""
    q = (q or "").strip()
    if len(q) < 2:
        return {"results": [], "count": 0}
    rx = {"$regex": re.escape(q), "$options": "i"}
    results = []

    equipments = await db.equipments.find(
        {"$or": [{"reference": rx}, {"numero_serie": rx}, {"type": rx}, {"description": rx}]},
        {"_id": 0}).to_list(15)
    for e in equipments:
        results.append({
            "category": "equipment", "label_category": "Équipement",
            "id": e["id"], "label": e.get("reference") or e.get("type"),
            "sublabel": f"{e.get('type', '')} · S/N {e.get('numero_serie', '')}".strip(" ·"),
        })

    subs = await db.subequipments.find(
        {"$or": [{"nom": rx}, {"reference": rx}, {"numero_serie": rx}]},
        {"_id": 0}).to_list(15)
    for s in subs:
        results.append({
            "category": "subequipment", "label_category": "Sous-équipement",
            "id": s["id"], "label": s.get("nom"),
            "sublabel": f"Réf {s.get('reference', '')}".strip(),
        })

    work_orders = await db.work_orders.find(
        {"$or": [{"titre": rx}, {"description": rx}, {"technicien_assigne": rx}]},
        {"_id": 0}).to_list(15)
    for w in work_orders:
        results.append({
            "category": "work_order", "label_category": "Maintenance",
            "id": w["id"], "label": w.get("titre"),
            "sublabel": f"{w.get('type_maintenance', '')} · {w.get('statut', '')} · {w.get('date_planifiee', '')}".strip(" ·"),
        })

    interventions = await db.interventions.find(
        {"$or": [{"actions_realisees": rx}, {"technicien": rx}, {"observations": rx}]},
        {"_id": 0}).to_list(15)
    for it in interventions:
        results.append({
            "category": "intervention", "label_category": "Intervention",
            "id": it["id"], "label": it.get("actions_realisees") or "Intervention",
            "sublabel": f"{it.get('technicien', '')} · {it.get('date_intervention', '')}".strip(" ·"),
        })

    inspections = await db.inspections.find(
        {"$or": [{"titre": rx}, {"organisme_certificateur": rx}]},
        {"_id": 0}).to_list(15)
    for insp in inspections:
        results.append({
            "category": "inspection", "label_category": "Contrôle réglementaire",
            "id": insp["id"], "label": insp.get("titre"),
            "sublabel": f"{insp.get('type_controle', '')} · échéance {insp.get('date_validite', '')}".strip(" ·"),
        })

    return {"results": results, "count": len(results)}


# ==================== EXPORT ROUTES ====================

@api_router.get("/export/xlsx/{collection}")
async def export_collection_xlsx(collection: str, current_user: dict = Depends(get_current_user)):
    # Check permission
    if not can_export(current_user):
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs et techniciens")
    
    sheet_names = {
        "equipments": "Equipements",
        "work_orders": "Ordres_Travail",
        "interventions": "Interventions",
        "inspections": "Controles",
        "spare_parts": "Pieces_Detachees",
    }
    if collection not in sheet_names:
        raise HTTPException(status_code=400, detail="Collection invalide")
    
    data = await db[collection].find({}, {"_id": 0}).to_list(10000)
    
    if not data:
        raise HTTPException(status_code=404, detail="Aucune donnée à exporter")
    
    output = io.BytesIO()
    df = pd.DataFrame(data)
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_names[collection], index=False)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={collection}.xlsx"}
    )

@api_router.get("/export/sql")
async def export_sql(current_user: dict = Depends(get_current_user)):
    # Check permission
    if not can_export(current_user):
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs et techniciens")
    
    collections = ["caisson", "equipments", "work_orders", "interventions", "inspections", "spare_parts", "users"]
    
    sql_output = []
    sql_output.append("-- HyperbareManager Database Export")
    sql_output.append(f"-- Generated: {datetime.now(timezone.utc).isoformat()}")
    sql_output.append("")
    
    for coll_name in collections:
        data = await db[coll_name].find({}, {"_id": 0}).to_list(10000)
        if data:
            # Create table statement
            sample = data[0]
            columns = ", ".join([f"{k} TEXT" for k in sample.keys()])
            sql_output.append(f"CREATE TABLE IF NOT EXISTS {coll_name} ({columns});")
            sql_output.append("")
            
            # Insert statements
            for row in data:
                cols = ", ".join(row.keys())
                vals = ", ".join([f"'{str(v).replace(chr(39), chr(39)+chr(39))}'" for v in row.values()])
                sql_output.append(f"INSERT INTO {coll_name} ({cols}) VALUES ({vals});")
            sql_output.append("")
    
    output = "\n".join(sql_output)
    
    return StreamingResponse(
        iter([output]),
        media_type="application/sql",
        headers={"Content-Disposition": "attachment; filename=hyperbaremanager_export.sql"}
    )

@api_router.get("/export/json")
async def export_json(current_user: dict = Depends(get_current_user)):
    # Check permission
    if not can_export(current_user):
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs et techniciens")
    
    collections = ["caisson", "equipments", "work_orders", "interventions", "inspections", "spare_parts"]
    
    export_data = {}
    for coll_name in collections:
        data = await db[coll_name].find({}, {"_id": 0}).to_list(10000)
        export_data[coll_name] = data
    
    output = json.dumps(export_data, indent=2, ensure_ascii=False, default=str)
    
    return StreamingResponse(
        iter([output]),
        media_type="application/json",
        headers={"Content-Disposition": "attachment; filename=hyperbaremanager_export.json"}
    )

@api_router.get("/export/excel")
async def export_excel(current_user: dict = Depends(get_current_user)):
    """Export all data to Excel for audit purposes"""
    if not can_export(current_user):
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs et techniciens")
    
    # Create Excel file in memory
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Equipments
        equipments = await db.equipments.find({}, {"_id": 0}).to_list(10000)
        if equipments:
            df = pd.DataFrame(equipments)
            df.to_excel(writer, sheet_name='Equipements', index=False)
        
        # Work Orders
        work_orders = await db.work_orders.find({}, {"_id": 0}).to_list(10000)
        if work_orders:
            df = pd.DataFrame(work_orders)
            df.to_excel(writer, sheet_name='Ordres_Travail', index=False)
        
        # Interventions
        interventions = await db.interventions.find({}, {"_id": 0}).to_list(10000)
        if interventions:
            df = pd.DataFrame(interventions)
            df.to_excel(writer, sheet_name='Interventions', index=False)
        
        # Inspections
        inspections = await db.inspections.find({}, {"_id": 0}).to_list(10000)
        if inspections:
            df = pd.DataFrame(inspections)
            df.to_excel(writer, sheet_name='Inspections', index=False)
        
        # Spare Parts
        spare_parts = await db.spare_parts.find({}, {"_id": 0}).to_list(10000)
        if spare_parts:
            df = pd.DataFrame(spare_parts)
            df.to_excel(writer, sheet_name='Pieces_Detachees', index=False)
        
        # Gas Cylinders
        gas_cylinders = await db.gas_cylinders.find({}, {"_id": 0}).to_list(10000)
        if gas_cylinders:
            df = pd.DataFrame(gas_cylinders)
            df.to_excel(writer, sheet_name='Bouteilles_Gaz', index=False)
        
        # Contractors
        contractors = await db.contractors.find({}, {"_id": 0}).to_list(10000)
        if contractors:
            df = pd.DataFrame(contractors)
            df.to_excel(writer, sheet_name='Prestataires', index=False)
        
        # Contracts
        contracts = await db.contracts.find({}, {"_id": 0}).to_list(10000)
        if contracts:
            df = pd.DataFrame(contracts)
            df.to_excel(writer, sheet_name='Contrats', index=False)
        
        # Budget
        budget = await db.budget.find({}, {"_id": 0}).to_list(10000)
        if budget:
            df = pd.DataFrame(budget)
            df.to_excel(writer, sheet_name='Budget', index=False)
        
        # Control Reports
        control_reports = await db.control_reports.find({}, {"_id": 0}).to_list(10000)
        if control_reports:
            df = pd.DataFrame(control_reports)
            df.to_excel(writer, sheet_name='PV_Controle', index=False)
        
        # Documents
        documents = await db.documents.find({}, {"_id": 0}).to_list(10000)
        if documents:
            df = pd.DataFrame(documents)
            df.to_excel(writer, sheet_name='Documents', index=False)
    
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=hyperbaremanager_audit_{datetime.now().strftime('%Y%m%d')}.xlsx"}
    )

# ==================== FILE UPLOAD ROUTES ====================

ALLOWED_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".gif", ".webp"}
ALLOWED_DOC_EXTENSIONS = {".pdf"}
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB

def get_file_extension(filename: str) -> str:
    return Path(filename).suffix.lower()

@api_router.post("/equipments/{equipment_id}/photos")
async def upload_equipment_photo(
    equipment_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Upload a photo for an equipment"""
    # Verify equipment exists
    equipment = await db.equipments.find_one({"id": equipment_id})
    if not equipment:
        raise HTTPException(status_code=404, detail="Équipement non trouvé")
    
    # Validate file extension
    ext = get_file_extension(file.filename)
    if ext not in ALLOWED_IMAGE_EXTENSIONS:
        raise HTTPException(status_code=400, detail=f"Format non supporté. Formats acceptés: {', '.join(ALLOWED_IMAGE_EXTENSIONS)}")
    
    # Generate unique filename
    unique_filename = f"{uuid.uuid4()}{ext}"
    file_path = UPLOADS_DIR / "equipments" / unique_filename
    
    # Save file
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    # Update equipment in database
    photo_url = f"/api/uploads/equipments/{unique_filename}"
    await db.equipments.update_one(
        {"id": equipment_id},
        {"$push": {"photos": photo_url}}
    )
    
    return {"filename": file.filename, "url": photo_url}

@api_router.post("/equipments/{equipment_id}/documents")
async def upload_equipment_document(
    equipment_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Upload a PDF document for an equipment"""
    # Verify equipment exists
    equipment = await db.equipments.find_one({"id": equipment_id})
    if not equipment:
        raise HTTPException(status_code=404, detail="Équipement non trouvé")
    
    # Validate file extension
    ext = get_file_extension(file.filename)
    if ext not in ALLOWED_DOC_EXTENSIONS:
        raise HTTPException(status_code=400, detail="Seuls les fichiers PDF sont acceptés")
    
    # Generate unique filename
    unique_filename = f"{uuid.uuid4()}{ext}"
    file_path = UPLOADS_DIR / "equipments" / unique_filename
    
    # Save file
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    # Update equipment in database
    doc_url = f"/api/uploads/equipments/{unique_filename}"
    doc_info = {
        "filename": file.filename,
        "url": doc_url,
        "uploaded_at": datetime.now(timezone.utc).isoformat()
    }
    await db.equipments.update_one(
        {"id": equipment_id},
        {"$push": {"documents": doc_info}}
    )
    
    return doc_info

@api_router.delete("/equipments/{equipment_id}/photos")
async def delete_equipment_photo(
    equipment_id: str,
    photo_url: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete a photo from an equipment"""
    # Remove from database
    await db.equipments.update_one(
        {"id": equipment_id},
        {"$pull": {"photos": photo_url}}
    )
    
    # Delete file
    filename = photo_url.split("/")[-1]
    file_path = UPLOADS_DIR / "equipments" / filename
    if file_path.exists():
        file_path.unlink()
    
    return {"message": "Photo supprimée"}

@api_router.delete("/equipments/{equipment_id}/documents")
async def delete_equipment_document(
    equipment_id: str,
    doc_url: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete a document from an equipment"""
    # Remove from database
    await db.equipments.update_one(
        {"id": equipment_id},
        {"$pull": {"documents": {"url": doc_url}}}
    )
    
    # Delete file
    filename = doc_url.split("/")[-1]
    file_path = UPLOADS_DIR / "equipments" / filename
    if file_path.exists():
        file_path.unlink()
    
    return {"message": "Document supprimé"}

@api_router.post("/inspections/{inspection_id}/procedures")
async def upload_inspection_procedure(
    inspection_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Upload a procedure PDF for an inspection"""
    # Verify inspection exists
    inspection = await db.inspections.find_one({"id": inspection_id})
    if not inspection:
        raise HTTPException(status_code=404, detail="Contrôle non trouvé")
    
    # Validate file extension
    ext = get_file_extension(file.filename)
    if ext not in ALLOWED_DOC_EXTENSIONS:
        raise HTTPException(status_code=400, detail="Seuls les fichiers PDF sont acceptés")
    
    # Generate unique filename
    unique_filename = f"{uuid.uuid4()}{ext}"
    file_path = UPLOADS_DIR / "inspections" / unique_filename
    
    # Save file
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    # Update inspection in database
    doc_url = f"/api/uploads/inspections/{unique_filename}"
    doc_info = {
        "filename": file.filename,
        "url": doc_url,
        "uploaded_at": datetime.now(timezone.utc).isoformat()
    }
    await db.inspections.update_one(
        {"id": inspection_id},
        {"$push": {"procedure_documents": doc_info}}
    )
    
    return doc_info

@api_router.delete("/inspections/{inspection_id}/procedures")
async def delete_inspection_procedure(
    inspection_id: str,
    doc_url: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete a procedure from an inspection"""
    # Remove from database
    await db.inspections.update_one(
        {"id": inspection_id},
        {"$pull": {"procedure_documents": {"url": doc_url}}}
    )
    
    # Delete file
    filename = doc_url.split("/")[-1]
    file_path = UPLOADS_DIR / "inspections" / filename
    if file_path.exists():
        file_path.unlink()
    
    return {"message": "Procédure supprimée"}

# Serve uploaded files
@api_router.get("/uploads/{folder}/{filename}")
async def get_uploaded_file(folder: str, filename: str):
    """Serve uploaded files"""
    if folder not in ["equipments", "inspections", "subequipments", "spareparts", "workorders", "interventions"]:
        raise HTTPException(status_code=404, detail="Dossier non trouvé")
    
    file_path = UPLOADS_DIR / folder / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Fichier non trouvé")
    
    # Determine content type
    ext = get_file_extension(filename)
    content_types = {
        ".pdf": "application/pdf",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png": "image/png",
        ".gif": "image/gif",
        ".webp": "image/webp"
    }
    content_type = content_types.get(ext, "application/octet-stream")
    
    return FileResponse(file_path, media_type=content_type)

# ==================== MAINTENANCE REPORT ====================

@api_router.get("/reports/maintenance")
async def get_maintenance_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Generate a maintenance report"""
    query = {}
    
    # Filter by date range if provided
    if start_date or end_date:
        date_filter = {}
        if start_date:
            date_filter["$gte"] = start_date
        if end_date:
            date_filter["$lte"] = end_date
        query["date_intervention"] = date_filter
    
    # Get interventions
    interventions = await db.interventions.find(query, {"_id": 0}).to_list(10000)
    
    # Get work orders for reference
    work_orders = await db.work_orders.find({}, {"_id": 0}).to_list(10000)
    wo_dict = {wo["id"]: wo for wo in work_orders}
    
    # Get equipments for reference
    equipments = await db.equipments.find({}, {"_id": 0}).to_list(10000)
    eq_dict = {eq["id"]: eq for eq in equipments}
    
    # Build report
    report = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "period": {
            "start": start_date or "Début",
            "end": end_date or "Aujourd'hui"
        },
        "summary": {
            "total_interventions": len(interventions),
            "total_duration_minutes": sum(i.get("duree_minutes", 0) or 0 for i in interventions),
            "preventive_count": 0,
            "corrective_count": 0
        },
        "interventions": []
    }
    
    for intervention in interventions:
        wo = wo_dict.get(intervention.get("work_order_id"), {})
        eq = eq_dict.get(wo.get("equipment_id"), {})
        
        if wo.get("type_maintenance") == "preventive":
            report["summary"]["preventive_count"] += 1
        else:
            report["summary"]["corrective_count"] += 1
        
        report["interventions"].append({
            "date": intervention.get("date_intervention"),
            "technicien": intervention.get("technicien"),
            "ordre_travail": wo.get("titre", "N/A"),
            "type_maintenance": wo.get("type_maintenance", "N/A"),
            "equipement": f"{eq.get('type', 'Caisson')} - {eq.get('reference', 'N/A')}",
            "actions": intervention.get("actions_realisees"),
            "duree_minutes": intervention.get("duree_minutes"),
            "observations": intervention.get("observations"),
            "pieces_utilisees": intervention.get("pieces_utilisees", [])
        })
    
    return report

@api_router.get("/reports/maintenance/csv")
async def export_maintenance_report_csv(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Export maintenance report as CSV"""
    # Check permission
    if not can_export(current_user):
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs et techniciens")
    
    report = await get_maintenance_report(start_date, end_date, current_user)
    
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')
    
    # Header
    writer.writerow([
        "Date", "Technicien", "Ordre de travail", "Type maintenance",
        "Équipement", "Actions réalisées", "Durée (min)", "Observations"
    ])
    
    # Data
    for i in report["interventions"]:
        writer.writerow([
            i["date"],
            i["technicien"],
            i["ordre_travail"],
            i["type_maintenance"],
            i["equipement"],
            i["actions"],
            i["duree_minutes"] or "",
            i["observations"] or ""
        ])
    
    output.seek(0)
    filename = f"rapport_maintenance_{start_date or 'debut'}_{end_date or 'fin'}.csv"
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/reports/statistics")
async def get_statistics_report(current_user: dict = Depends(get_current_user)):
    """Get comprehensive statistics report"""
    if not can_export(current_user):
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs et techniciens")
    
    today = datetime.now(timezone.utc).date()
    
    # Equipment stats
    equipments = await db.equipments.find({}, {"_id": 0}).to_list(1000)
    equipment_by_type = {}
    equipment_by_status = {"en_service": 0, "maintenance": 0, "hors_service": 0}
    equipment_by_criticite = {"critique": 0, "haute": 0, "normale": 0, "basse": 0}
    
    for eq in equipments:
        eq_type = eq.get("type", "Autre")
        equipment_by_type[eq_type] = equipment_by_type.get(eq_type, 0) + 1
        equipment_by_status[eq.get("statut", "en_service")] = equipment_by_status.get(eq.get("statut", "en_service"), 0) + 1
        equipment_by_criticite[eq.get("criticite", "normale")] = equipment_by_criticite.get(eq.get("criticite", "normale"), 0) + 1
    
    # Work order stats
    work_orders = await db.work_orders.find({}, {"_id": 0}).to_list(1000)
    wo_by_status = {"planifiee": 0, "en_cours": 0, "terminee": 0, "annulee": 0}
    wo_by_type = {"preventive": 0, "corrective": 0}
    overdue_count = 0
    
    for wo in work_orders:
        wo_by_status[wo.get("statut", "planifiee")] = wo_by_status.get(wo.get("statut", "planifiee"), 0) + 1
        wo_by_type[wo.get("type_maintenance", "corrective")] = wo_by_type.get(wo.get("type_maintenance", "corrective"), 0) + 1
        try:
            planned_date = datetime.strptime(wo["date_planifiee"], "%Y-%m-%d").date()
            if planned_date < today and wo.get("statut") in ["planifiee", "en_cours"]:
                overdue_count += 1
        except:
            pass
    
    # Interventions stats
    interventions = await db.interventions.find({}, {"_id": 0}).to_list(1000)
    total_duration = sum(i.get("duree_minutes", 0) or 0 for i in interventions)
    
    # Inspections stats
    inspections = await db.inspections.find({}, {"_id": 0}).to_list(1000)
    expired_inspections = 0
    upcoming_inspections = 0
    
    for insp in inspections:
        try:
            validity_date = datetime.strptime(insp.get("date_validite", ""), "%Y-%m-%d").date()
            if validity_date < today:
                expired_inspections += 1
            elif (validity_date - today).days <= 30:
                upcoming_inspections += 1
        except:
            pass
    
    # Spare parts stats
    spare_parts = await db.spare_parts.find({}, {"_id": 0}).to_list(1000)
    low_stock_count = sum(1 for p in spare_parts if p.get("quantite_stock", 0) <= p.get("seuil_minimum", 1))
    total_stock_value = sum((p.get("quantite_stock", 0) * (p.get("prix_unitaire", 0) or 0)) for p in spare_parts)
    
    return {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "equipments": {
            "total": len(equipments),
            "by_type": equipment_by_type,
            "by_status": equipment_by_status,
            "by_criticite": equipment_by_criticite
        },
        "work_orders": {
            "total": len(work_orders),
            "by_status": wo_by_status,
            "by_type": wo_by_type,
            "overdue": overdue_count
        },
        "interventions": {
            "total": len(interventions),
            "total_duration_minutes": total_duration,
            "average_duration_minutes": round(total_duration / len(interventions), 1) if interventions else 0
        },
        "inspections": {
            "total": len(inspections),
            "expired": expired_inspections,
            "upcoming_30_days": upcoming_inspections
        },
        "spare_parts": {
            "total": len(spare_parts),
            "low_stock": low_stock_count,
            "total_stock_value": round(total_stock_value, 2)
        }
    }

@api_router.get("/reports/statistics/csv")
async def export_statistics_csv(current_user: dict = Depends(get_current_user)):
    """Export statistics as CSV"""
    if not can_export(current_user):
        raise HTTPException(status_code=403, detail="Accès réservé aux administrateurs et techniciens")
    
    stats = await get_statistics_report(current_user)
    
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')
    
    # Write statistics
    writer.writerow(["Rapport de statistiques HyperbareManager"])
    writer.writerow(["Généré le", stats["generated_at"]])
    writer.writerow([])
    
    writer.writerow(["=== ÉQUIPEMENTS ==="])
    writer.writerow(["Total", stats["equipments"]["total"]])
    writer.writerow(["Par statut"])
    for k, v in stats["equipments"]["by_status"].items():
        writer.writerow(["", k, v])
    writer.writerow(["Par type"])
    for k, v in stats["equipments"]["by_type"].items():
        writer.writerow(["", k, v])
    writer.writerow([])
    
    writer.writerow(["=== ORDRES DE TRAVAIL ==="])
    writer.writerow(["Total", stats["work_orders"]["total"]])
    writer.writerow(["En retard", stats["work_orders"]["overdue"]])
    writer.writerow(["Par statut"])
    for k, v in stats["work_orders"]["by_status"].items():
        writer.writerow(["", k, v])
    writer.writerow([])
    
    writer.writerow(["=== INTERVENTIONS ==="])
    writer.writerow(["Total", stats["interventions"]["total"]])
    writer.writerow(["Durée totale (min)", stats["interventions"]["total_duration_minutes"]])
    writer.writerow(["Durée moyenne (min)", stats["interventions"]["average_duration_minutes"]])
    writer.writerow([])
    
    writer.writerow(["=== CONTRÔLES RÉGLEMENTAIRES ==="])
    writer.writerow(["Total", stats["inspections"]["total"]])
    writer.writerow(["Expirés", stats["inspections"]["expired"]])
    writer.writerow(["À renouveler (30j)", stats["inspections"]["upcoming_30_days"]])
    writer.writerow([])
    
    writer.writerow(["=== PIÈCES DÉTACHÉES ==="])
    writer.writerow(["Total", stats["spare_parts"]["total"]])
    writer.writerow(["Stock bas", stats["spare_parts"]["low_stock"]])
    writer.writerow(["Valeur totale stock (€)", stats["spare_parts"]["total_stock_value"]])
    
    output.seek(0)
    filename = f"statistiques_hyperbaremanager_{datetime.now(timezone.utc).strftime('%Y%m%d')}.csv"
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ==================== PDF REPORTS ====================

def create_pdf_styles():
    """Create custom PDF styles"""
    styles = getSampleStyleSheet()
    
    # Title style
    styles.add(ParagraphStyle(
        name='PDFTitle',
        parent=styles['Heading1'],
        fontSize=20,
        spaceAfter=20,
        textColor=colors.HexColor('#005F73'),
        alignment=TA_CENTER
    ))
    
    # Subtitle style
    styles.add(ParagraphStyle(
        name='PDFSubtitle',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=10,
        textColor=colors.HexColor('#005F73')
    ))
    
    # Section header
    styles.add(ParagraphStyle(
        name='SectionHeader',
        parent=styles['Heading3'],
        fontSize=12,
        spaceBefore=15,
        spaceAfter=8,
        textColor=colors.HexColor('#0A9396'),
        borderPadding=5
    ))
    
    # Normal text
    styles.add(ParagraphStyle(
        name='PDFNormal',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=6
    ))
    
    # Small text
    styles.add(ParagraphStyle(
        name='PDFSmall',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.gray
    ))
    
    return styles

def create_pdf_header(title: str, subtitle: str = None):
    """Create PDF header elements"""
    styles = create_pdf_styles()
    elements = []
    
    # Header with logo placeholder
    header_data = [[
        Paragraph(f"<b>HyperbareManager</b>", styles['PDFTitle']),
        Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", styles['PDFSmall'])
    ]]
    
    header_table = Table(header_data, colWidths=[12*cm, 6*cm])
    header_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 10))
    
    # Title
    elements.append(Paragraph(title, styles['PDFTitle']))
    
    if subtitle:
        elements.append(Paragraph(subtitle, styles['PDFSubtitle']))
    
    elements.append(Spacer(1, 20))
    
    return elements, styles

def create_table_style():
    """Create standard table style"""
    return TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#005F73')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E0E0E0')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
    ])

@api_router.get("/reports/pdf/statistics")
async def generate_statistics_pdf(current_user: dict = Depends(get_current_user)):
    """Generate statistics PDF report"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    
    elements, styles = create_pdf_header("Rapport de Statistiques", "Vue d'ensemble de l'installation")
    
    # Get data
    caisson = await db.caisson.find_one({}, {"_id": 0})
    equipments = await db.equipments.find({}, {"_id": 0}).to_list(1000)
    work_orders = await db.work_orders.find({}, {"_id": 0}).to_list(1000)
    interventions = await db.interventions.find({}, {"_id": 0}).to_list(1000)
    spare_parts = await db.spare_parts.find({}, {"_id": 0}).to_list(1000)
    
    # Caisson info
    if caisson:
        elements.append(Paragraph("Informations du Caisson", styles['SectionHeader']))
        caisson_data = [
            ["Nom", caisson.get("nom", "N/A")],
            ["Modèle", caisson.get("modele", "N/A")],
            ["Fabricant", caisson.get("fabricant", "N/A")],
            ["N° Série", caisson.get("numero_serie", "N/A")],
        ]
        t = Table(caisson_data, colWidths=[5*cm, 12*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F0F0F0')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E0E0E0')),
            ('PADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 15))
    
    # Equipment stats
    elements.append(Paragraph("Statistiques des Équipements", styles['SectionHeader']))
    eq_en_service = len([e for e in equipments if e.get('statut') == 'en_service'])
    eq_maintenance = len([e for e in equipments if e.get('statut') == 'maintenance'])
    eq_hors_service = len([e for e in equipments if e.get('statut') == 'hors_service'])
    
    eq_data = [
        ["Statut", "Nombre"],
        ["En service", str(eq_en_service)],
        ["En maintenance", str(eq_maintenance)],
        ["Hors service", str(eq_hors_service)],
        ["Total", str(len(equipments))],
    ]
    t = Table(eq_data, colWidths=[10*cm, 4*cm])
    t.setStyle(create_table_style())
    elements.append(t)
    elements.append(Spacer(1, 15))
    
    # Work orders stats
    elements.append(Paragraph("Statistiques des Maintenances", styles['SectionHeader']))
    wo_planifiee = len([w for w in work_orders if w.get('statut') == 'planifiee'])
    wo_en_cours = len([w for w in work_orders if w.get('statut') == 'en_cours'])
    wo_terminee = len([w for w in work_orders if w.get('statut') == 'terminee'])
    
    wo_data = [
        ["Statut", "Nombre"],
        ["Planifiées", str(wo_planifiee)],
        ["En cours", str(wo_en_cours)],
        ["Terminées", str(wo_terminee)],
        ["Total", str(len(work_orders))],
    ]
    t = Table(wo_data, colWidths=[10*cm, 4*cm])
    t.setStyle(create_table_style())
    elements.append(t)
    elements.append(Spacer(1, 15))
    
    # Interventions stats
    elements.append(Paragraph("Statistiques des Interventions", styles['SectionHeader']))
    elements.append(Paragraph(f"<b>Total des interventions :</b> {len(interventions)}", styles['PDFNormal']))
    elements.append(Spacer(1, 15))
    
    # Spare parts stats
    elements.append(Paragraph("Stock de Pièces Détachées", styles['SectionHeader']))
    low_stock = len([p for p in spare_parts if (p.get('quantite_stock') or 0) <= (p.get('seuil_minimum') or 1)])
    total_value = sum((p.get('quantite_stock') or 0) * (p.get('prix_unitaire') or 0) for p in spare_parts)
    
    sp_data = [
        ["Indicateur", "Valeur"],
        ["Références en stock", str(len(spare_parts))],
        ["Alertes stock bas", str(low_stock)],
        ["Valeur totale du stock", f"{total_value:.2f} €"],
    ]
    t = Table(sp_data, colWidths=[10*cm, 4*cm])
    t.setStyle(create_table_style())
    elements.append(t)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"statistiques_hyperbaremanager_{datetime.now().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/reports/pdf/maintenance")
async def generate_maintenance_pdf(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Generate maintenance history PDF report"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    
    period = ""
    if start_date and end_date:
        period = f"Du {start_date} au {end_date}"
    
    elements, styles = create_pdf_header("Rapport de Maintenance", period)
    
    # Get work orders
    query = {}
    work_orders = await db.work_orders.find(query, {"_id": 0}).to_list(1000)
    
    # Filter by date if provided
    if start_date:
        work_orders = [w for w in work_orders if w.get('date_planifiee', '') >= start_date]
    if end_date:
        work_orders = [w for w in work_orders if w.get('date_planifiee', '') <= end_date]
    
    # Get equipment map
    equipments = await db.equipments.find({}, {"_id": 0}).to_list(1000)
    eq_map = {e['id']: e for e in equipments}
    
    # Summary
    elements.append(Paragraph("Résumé", styles['SectionHeader']))
    summary_data = [
        ["Type", "Nombre"],
        ["Maintenances préventives", str(len([w for w in work_orders if w.get('type_maintenance') == 'preventive']))],
        ["Maintenances curatives", str(len([w for w in work_orders if w.get('type_maintenance') == 'curative']))],
        ["Total", str(len(work_orders))],
    ]
    t = Table(summary_data, colWidths=[10*cm, 4*cm])
    t.setStyle(create_table_style())
    elements.append(t)
    elements.append(Spacer(1, 20))
    
    # Detail table
    elements.append(Paragraph("Détail des Maintenances", styles['SectionHeader']))
    
    if work_orders:
        detail_data = [["Titre", "Équipement", "Type", "Statut", "Date"]]
        for wo in work_orders[:50]:  # Limit to 50
            eq = eq_map.get(wo.get('equipment_id'), {})
            detail_data.append([
                wo.get('titre', 'N/A')[:30],
                eq.get('reference', 'Caisson')[:20],
                'Préventive' if wo.get('type_maintenance') == 'preventive' else 'Curative',
                wo.get('statut', 'N/A'),
                wo.get('date_planifiee', 'N/A')[:10] if wo.get('date_planifiee') else 'N/A'
            ])
        
        t = Table(detail_data, colWidths=[5*cm, 4*cm, 2.5*cm, 2.5*cm, 2.5*cm])
        t.setStyle(create_table_style())
        elements.append(t)
    else:
        elements.append(Paragraph("Aucune maintenance trouvée pour cette période.", styles['PDFNormal']))
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"rapport_maintenance_{datetime.now().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/reports/pdf/equipment/{equipment_id}")
async def generate_equipment_pdf(equipment_id: str, current_user: dict = Depends(get_current_user)):
    """Generate equipment detail PDF (fiche équipement)"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    
    # Get equipment
    equipment = await db.equipments.find_one({"id": equipment_id}, {"_id": 0})
    if not equipment:
        raise HTTPException(status_code=404, detail="Équipement non trouvé")
    
    elements, styles = create_pdf_header("Fiche Équipement", equipment.get('reference', 'N/A'))
    
    # Equipment info
    elements.append(Paragraph("Informations Générales", styles['SectionHeader']))
    
    info_data = [
        ["Type", equipment.get('type', 'N/A')],
        ["Référence", equipment.get('reference', 'N/A')],
        ["N° Série", equipment.get('numero_serie', 'N/A')],
        ["Statut", equipment.get('statut', 'N/A')],
        ["Criticité", equipment.get('criticite', 'N/A')],
        ["Date d'installation", equipment.get('date_installation', 'N/A')[:10] if equipment.get('date_installation') else 'N/A'],
    ]
    
    if (equipment.get('type') or '').lower() == 'compresseur':
        info_data.append(["Compteur horaire", f"{equipment.get('compteur_horaire', 0):,.0f} h"])
    
    t = Table(info_data, colWidths=[5*cm, 12*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F0F0F0')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E0E0E0')),
        ('PADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 20))
    
    # Description
    if equipment.get('description'):
        elements.append(Paragraph("Description", styles['SectionHeader']))
        elements.append(Paragraph(equipment.get('description'), styles['PDFNormal']))
        elements.append(Spacer(1, 15))
    
    # Get interventions rattachées directement à l'équipement
    work_orders = await db.work_orders.find({"equipment_id": equipment_id}, {"_id": 0}).to_list(200)
    interventions = await db.interventions.find({"equipment_id": equipment_id}, {"_id": 0}).sort("date_intervention", -1).to_list(1000)

    # Historique des réformes / changements de statut
    hist_statut = equipment.get('historique_statut') or []
    if hist_statut:
        elements.append(Paragraph("Historique des Réformes / Changements de Statut", styles['SectionHeader']))
        statut_data = [["Date", "Ancien", "Nouveau", "Motif", "Par"]]
        for h in reversed(hist_statut):
            statut_data.append([
                (h.get('date') or '')[:10],
                h.get('ancien_statut') or '-',
                h.get('nouveau_statut') or '-',
                (h.get('motif') or '-')[:22],
                (h.get('utilisateur') or '-')[:20],
            ])
        t = Table(statut_data, colWidths=[2.3*cm, 2.6*cm, 2.6*cm, 5*cm, 4.5*cm])
        t.setStyle(create_table_style())
        elements.append(t)
        elements.append(Spacer(1, 20))

    # Maintenance history
    elements.append(Paragraph("Historique des Maintenances", styles['SectionHeader']))
    
    if work_orders:
        maint_data = [["Titre", "Type", "Statut", "Date"]]
        for wo in work_orders[:20]:
            maint_data.append([
                wo.get('titre', 'N/A')[:35],
                'Préventive' if wo.get('type_maintenance') == 'preventive' else 'Curative',
                wo.get('statut', 'N/A'),
                wo.get('date_planifiee', 'N/A')[:10] if wo.get('date_planifiee') else 'N/A'
            ])
        
        t = Table(maint_data, colWidths=[7*cm, 3*cm, 3*cm, 3*cm])
        t.setStyle(create_table_style())
        elements.append(t)
    else:
        elements.append(Paragraph("Aucune maintenance enregistrée.", styles['PDFNormal']))
    
    elements.append(Spacer(1, 20))
    
    # Interventions history
    elements.append(Paragraph("Historique des Interventions", styles['SectionHeader']))

    if interventions:
        int_data = [["Date", "Type", "Actions réalisées", "Technicien"]]
        for inter in interventions[:60]:
            int_data.append([
                (inter.get('date_intervention') or 'N/A')[:10],
                'Préventive' if inter.get('type_intervention') == 'preventive' else 'Curative',
                (inter.get('actions_realisees') or 'N/A')[:45],
                (inter.get('technicien') or 'N/A')[:18],
            ])

        t = Table(int_data, colWidths=[2.3*cm, 2.2*cm, 8*cm, 3.5*cm])
        t.setStyle(create_table_style())
        elements.append(t)
        if len(interventions) > 60:
            elements.append(Spacer(1, 8))
            elements.append(Paragraph(
                f"... et {len(interventions) - 60} autre(s) intervention(s). Voir l'application pour l'historique complet.",
                styles['PDFNormal']))
    else:
        elements.append(Paragraph("Aucune intervention enregistrée.", styles['PDFNormal']))
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"fiche_equipement_{equipment.get('reference', 'unknown')}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/reports/pdf/interventions")
async def generate_interventions_pdf(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Generate interventions PDF report"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    
    period = ""
    if start_date and end_date:
        period = f"Du {start_date} au {end_date}"
    
    elements, styles = create_pdf_header("Rapport des Interventions", period)
    
    # Get data
    interventions = await db.interventions.find({}, {"_id": 0}).to_list(1000)
    
    # Filter by date
    if start_date:
        interventions = [i for i in interventions if i.get('date_realisation', '') >= start_date]
    if end_date:
        interventions = [i for i in interventions if i.get('date_realisation', '') <= end_date]
    
    # Get maps
    users = await db.users.find({}, {"_id": 0}).to_list(100)
    user_map = {u['id']: f"{u.get('prenom', '')} {u.get('nom', '')}" for u in users}
    
    work_orders = await db.work_orders.find({}, {"_id": 0}).to_list(1000)
    wo_map = {w['id']: w for w in work_orders}
    
    spare_parts = await db.spare_parts.find({}, {"_id": 0}).to_list(1000)
    sp_map = {s['id']: s for s in spare_parts}
    
    # Summary
    elements.append(Paragraph("Résumé", styles['SectionHeader']))
    elements.append(Paragraph(f"<b>Nombre total d'interventions :</b> {len(interventions)}", styles['PDFNormal']))
    
    # Count parts used
    total_parts = sum(len(i.get('pieces_utilisees', [])) for i in interventions)
    elements.append(Paragraph(f"<b>Pièces utilisées :</b> {total_parts}", styles['PDFNormal']))
    elements.append(Spacer(1, 20))
    
    # Detail table
    elements.append(Paragraph("Détail des Interventions", styles['SectionHeader']))
    
    if interventions:
        detail_data = [["Date", "Maintenance", "Technicien", "Pièces"]]
        for inter in interventions[:50]:
            wo = wo_map.get(inter.get('work_order_id'), {})
            pieces = inter.get('pieces_utilisees', [])
            pieces_str = ", ".join([sp_map.get(p.get('spare_part_id'), {}).get('nom', 'N/A')[:15] for p in pieces[:3]])
            if len(pieces) > 3:
                pieces_str += "..."
            
            detail_data.append([
                inter.get('date_realisation', 'N/A')[:10] if inter.get('date_realisation') else 'N/A',
                wo.get('titre', 'N/A')[:25],
                user_map.get(inter.get('technicien_id'), 'N/A')[:20],
                pieces_str or 'Aucune'
            ])
        
        t = Table(detail_data, colWidths=[3*cm, 5.5*cm, 4*cm, 4*cm])
        t.setStyle(create_table_style())
        elements.append(t)
    else:
        elements.append(Paragraph("Aucune intervention trouvée pour cette période.", styles['PDFNormal']))
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"rapport_interventions_{datetime.now().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/reports/pdf/planning")
async def generate_planning_pdf(current_user: dict = Depends(get_current_user)):
    """Generate maintenance planning PDF (52 weeks)"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    
    elements, styles = create_pdf_header("Planning de Maintenance", "52 prochaines semaines")
    
    # Get upcoming maintenances
    today = datetime.now(timezone.utc)
    end_date = today + timedelta(weeks=52)
    
    work_orders = await db.work_orders.find({
        "statut": {"$in": ["planifiee", "en_cours"]},
        "date_planifiee": {"$ne": None}
    }, {"_id": 0}).to_list(1000)
    
    # Filter by date range
    upcoming = []
    for wo in work_orders:
        try:
            date_str = wo.get('date_planifiee', '')
            if date_str:
                date_obj = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                if today <= date_obj <= end_date:
                    wo['date_obj'] = date_obj
                    upcoming.append(wo)
        except:
            pass
    
    # Sort by date
    upcoming.sort(key=lambda x: x.get('date_obj', today))
    
    # Get equipment map
    equipments = await db.equipments.find({}, {"_id": 0}).to_list(1000)
    eq_map = {e['id']: e for e in equipments}
    
    # Summary by month
    elements.append(Paragraph("Résumé par Mois", styles['SectionHeader']))
    
    months = {}
    for wo in upcoming:
        month_key = wo['date_obj'].strftime('%Y-%m')
        month_name = wo['date_obj'].strftime('%B %Y')
        if month_key not in months:
            months[month_key] = {'name': month_name, 'count': 0}
        months[month_key]['count'] += 1
    
    if months:
        month_data = [["Mois", "Maintenances planifiées"]]
        for key in sorted(months.keys()):
            month_data.append([months[key]['name'], str(months[key]['count'])])
        
        t = Table(month_data, colWidths=[10*cm, 5*cm])
        t.setStyle(create_table_style())
        elements.append(t)
    
    elements.append(Spacer(1, 20))
    
    # Detailed planning
    elements.append(Paragraph("Planning Détaillé", styles['SectionHeader']))
    
    if upcoming:
        plan_data = [["Date", "Titre", "Équipement", "Priorité"]]
        for wo in upcoming[:100]:
            eq = eq_map.get(wo.get('equipment_id'), {})
            plan_data.append([
                wo['date_obj'].strftime('%d/%m/%Y'),
                wo.get('titre', 'N/A')[:30],
                eq.get('reference', 'Caisson')[:20],
                wo.get('priorite', 'normale')
            ])
        
        t = Table(plan_data, colWidths=[3*cm, 6*cm, 4.5*cm, 3*cm])
        t.setStyle(create_table_style())
        elements.append(t)
    else:
        elements.append(Paragraph("Aucune maintenance planifiée pour les 52 prochaines semaines.", styles['PDFNormal']))
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"planning_maintenance_{datetime.now().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ==================== HEALTH CHECK ====================

@api_router.get("/")
async def root():
    return {"message": "HyperbareManager API", "version": "1.0.0"}

@api_router.get("/health")
async def health():
    return {"status": "healthy"}

# ==================== EMAIL ALERTS ====================

@api_router.post("/alerts/check")
async def check_and_send_alerts(admin: dict = Depends(require_admin)):
    """Check for alerts and send email notifications"""
    alerts_sent = {
        "maintenance_reminders": 0,
        "maintenance_overdue": 0,
        "low_stock": 0,
        "hour_counter": 0
    }
    
    # Get admin email for notifications
    admin_email = ADMIN_EMAIL or admin.get("email")
    if not admin_email:
        raise HTTPException(status_code=400, detail="Email admin non configuré")
    
    today = datetime.now(timezone.utc).date()
    
    # 1. Check maintenances coming up (30 days) and overdue
    maintenances = await db.work_orders.find({
        "statut": "planifiee",
        "date_planifiee": {"$ne": None}
    }, {"_id": 0}).to_list(1000)
    
    equipments_map = {}
    equipments = await db.equipments.find({}, {"_id": 0}).to_list(1000)
    for eq in equipments:
        equipments_map[eq["id"]] = eq
    
    for wo in maintenances:
        try:
            date_planifiee = datetime.fromisoformat(wo["date_planifiee"].replace("Z", "+00:00")).date() if isinstance(wo["date_planifiee"], str) else wo["date_planifiee"]
            if isinstance(date_planifiee, str):
                date_planifiee = datetime.strptime(date_planifiee, "%Y-%m-%d").date()
            
            days_diff = (date_planifiee - today).days
            equipment = equipments_map.get(wo.get("equipment_id"), {})
            equipment_ref = equipment.get("reference", "Caisson entier")
            
            if days_diff < 0:
                # Overdue
                await send_maintenance_overdue_email(
                    admin_email,
                    wo["titre"],
                    equipment_ref,
                    wo["date_planifiee"],
                    abs(days_diff)
                )
                alerts_sent["maintenance_overdue"] += 1
            elif days_diff <= 30:
                # Coming up in 30 days
                await send_maintenance_reminder_email(
                    admin_email,
                    wo["titre"],
                    equipment_ref,
                    wo["date_planifiee"],
                    days_diff
                )
                alerts_sent["maintenance_reminders"] += 1
        except Exception as e:
            logging.error(f"Error processing maintenance alert: {e}")
    
    # 2. Check low stock
    spare_parts = await db.spare_parts.find({}, {"_id": 0}).to_list(1000)
    for part in spare_parts:
        if part.get("quantite_stock", 0) <= part.get("seuil_minimum", 1):
            await send_low_stock_email(
                admin_email,
                part["nom"],
                part.get("reference_fabricant", "N/A"),
                part.get("quantite_stock", 0),
                part.get("seuil_minimum", 1)
            )
            alerts_sent["low_stock"] += 1
    
    # 3. Check hour counter alerts for compressors
    hour_maintenances = await db.work_orders.find({
        "statut": "planifiee",
        "periodicite_heures": {"$ne": None},
        "compteur_declenchement": {"$ne": None}
    }, {"_id": 0}).to_list(1000)
    
    for wo in hour_maintenances:
        equipment = equipments_map.get(wo.get("equipment_id"), {})
        if (equipment.get("type") or "").lower() == "compresseur":
            current_hours = equipment.get("compteur_horaire", 0) or 0
            threshold = wo.get("compteur_declenchement", 0) or 0
            if current_hours >= threshold:
                await send_hour_counter_alert_email(
                    admin_email,
                    equipment.get("reference", "Compresseur"),
                    current_hours,
                    threshold,
                    wo["titre"]
                )
                alerts_sent["hour_counter"] += 1
    
    return {
        "message": "Vérification des alertes terminée",
        "alerts_sent": alerts_sent,
        "total": sum(alerts_sent.values())
    }

@api_router.post("/alerts/test")
async def test_email(admin: dict = Depends(require_admin)):
    """Send a test email to verify configuration"""
    admin_email = ADMIN_EMAIL or admin.get("email")
    if not admin_email:
        raise HTTPException(status_code=400, detail="Email admin non configuré")
    
    content = """
    <p>Ceci est un email de test pour vérifier la configuration des notifications.</p>
    <p>Si vous recevez cet email, la configuration est correcte ! ✅</p>
    """
    
    success = await send_email(
        admin_email,
        "🧪 Test notification - HyperbareManager",
        email_template("Test de Configuration", content)
    )
    
    if success:
        return {"message": f"Email de test envoyé à {admin_email}"}
    else:
        raise HTTPException(status_code=500, detail="Échec de l'envoi de l'email")

# ==================== CONTRACTORS (PRESTATAIRES) ROUTES ====================

@api_router.get("/contractors", response_model=List[dict])
async def get_contractors(current_user: dict = Depends(get_current_user)):
    """Get all contractors/suppliers"""
    contractors = await db.contractors.find({}, {"_id": 0}).to_list(1000)
    return contractors

@api_router.get("/contractors/{contractor_id}")
async def get_contractor(contractor_id: str, current_user: dict = Depends(get_current_user)):
    """Get a single contractor"""
    contractor = await db.contractors.find_one({"id": contractor_id}, {"_id": 0})
    if not contractor:
        raise HTTPException(status_code=404, detail="Prestataire non trouvé")
    return contractor

@api_router.post("/contractors", response_model=dict)
async def create_contractor(contractor: ContractorCreate, current_user: dict = Depends(require_technicien_or_admin)):
    """Create a new contractor"""
    contractor_obj = Contractor(**contractor.model_dump())
    doc = contractor_obj.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    await db.contractors.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.put("/contractors/{contractor_id}")
async def update_contractor(contractor_id: str, contractor: ContractorCreate, current_user: dict = Depends(require_technicien_or_admin)):
    """Update a contractor"""
    existing = await db.contractors.find_one({"id": contractor_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Prestataire non trouvé")
    
    update_data = contractor.model_dump()
    await db.contractors.update_one({"id": contractor_id}, {"$set": update_data})
    updated = await db.contractors.find_one({"id": contractor_id}, {"_id": 0})
    return updated

@api_router.delete("/contractors/{contractor_id}")
async def delete_contractor(contractor_id: str, admin: dict = Depends(require_admin)):
    """Delete a contractor"""
    result = await db.contractors.delete_one({"id": contractor_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Prestataire non trouvé")
    return {"message": "Prestataire supprimé"}

# ==================== GAS CYLINDERS (BOUTEILLES DE GAZ) ROUTES ====================

@api_router.get("/gas-cylinders", response_model=List[dict])
async def get_gas_cylinders(
    type_gaz: Optional[str] = None,
    statut: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get all gas cylinders with optional filters"""
    query = {}
    if type_gaz:
        query["type_gaz"] = type_gaz
    if statut:
        query["statut"] = statut
    
    cylinders = await db.gas_cylinders.find(query, {"_id": 0}).to_list(1000)
    return cylinders

@api_router.get("/gas-cylinders/alerts")
async def get_gas_cylinder_alerts(current_user: dict = Depends(get_current_user)):
    """Get gas cylinders with expiring dates"""
    today = datetime.now(timezone.utc).date()
    alerts = {
        "gaz_expire": [],
        "epreuve_expire": [],
        "gaz_expire_30j": [],
        "epreuve_expire_90j": []
    }
    
    cylinders = await db.gas_cylinders.find({}, {"_id": 0}).to_list(1000)
    for cyl in cylinders:
        # Check gas expiration
        if cyl.get("date_expiration_gaz"):
            try:
                exp_date = datetime.strptime(cyl["date_expiration_gaz"], "%Y-%m-%d").date()
                days_left = (exp_date - today).days
                if days_left < 0:
                    alerts["gaz_expire"].append({**cyl, "jours_depasses": abs(days_left)})
                elif days_left <= 30:
                    alerts["gaz_expire_30j"].append({**cyl, "jours_restants": days_left})
            except:
                pass
        
        # Check hydraulic test expiration
        if cyl.get("date_prochaine_epreuve"):
            try:
                epr_date = datetime.strptime(cyl["date_prochaine_epreuve"], "%Y-%m-%d").date()
                days_left = (epr_date - today).days
                if days_left < 0:
                    alerts["epreuve_expire"].append({**cyl, "jours_depasses": abs(days_left)})
                elif days_left <= 90:
                    alerts["epreuve_expire_90j"].append({**cyl, "jours_restants": days_left})
            except:
                pass
    
    return alerts

@api_router.get("/gas-cylinders/{cylinder_id}")
async def get_gas_cylinder(cylinder_id: str, current_user: dict = Depends(get_current_user)):
    """Get a single gas cylinder"""
    cylinder = await db.gas_cylinders.find_one({"id": cylinder_id}, {"_id": 0})
    if not cylinder:
        raise HTTPException(status_code=404, detail="Bouteille non trouvée")
    return cylinder

@api_router.post("/gas-cylinders", response_model=dict)
async def create_gas_cylinder(cylinder: GasCylinderCreate, current_user: dict = Depends(require_technicien_or_admin)):
    """Create a new gas cylinder"""
    # Validate gas type
    if cylinder.type_gaz not in GAS_TYPES:
        raise HTTPException(status_code=400, detail=f"Type de gaz invalide. Types autorisés: {GAS_TYPES}")
    
    cylinder_obj = GasCylinder(**cylinder.model_dump())
    doc = cylinder_obj.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    await db.gas_cylinders.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.put("/gas-cylinders/{cylinder_id}")
async def update_gas_cylinder(cylinder_id: str, cylinder: GasCylinderCreate, current_user: dict = Depends(require_technicien_or_admin)):
    """Update a gas cylinder"""
    existing = await db.gas_cylinders.find_one({"id": cylinder_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Bouteille non trouvée")
    
    update_data = cylinder.model_dump()
    await db.gas_cylinders.update_one({"id": cylinder_id}, {"$set": update_data})
    updated = await db.gas_cylinders.find_one({"id": cylinder_id}, {"_id": 0})
    return updated

@api_router.post("/gas-cylinders/{cylinder_id}/refill")
async def refill_gas_cylinder(
    cylinder_id: str,
    date_remplissage: str = Form(...),
    date_expiration: str = Form(...),
    pression: Optional[float] = Form(None),
    agent: str = Form(...),
    observations: Optional[str] = Form(None),
    current_user: dict = Depends(require_technicien_or_admin)
):
    """Record a gas cylinder refill"""
    cylinder = await db.gas_cylinders.find_one({"id": cylinder_id})
    if not cylinder:
        raise HTTPException(status_code=404, detail="Bouteille non trouvée")
    
    refill_record = {
        "date": date_remplissage,
        "agent": agent,
        "pression": pression,
        "observations": observations,
        "recorded_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.gas_cylinders.update_one(
        {"id": cylinder_id},
        {
            "$set": {
                "date_remplissage": date_remplissage,
                "date_expiration_gaz": date_expiration,
                "statut": "pleine",
                "agent_responsable": agent
            },
            "$push": {"historique_remplissage": refill_record}
        }
    )
    
    updated = await db.gas_cylinders.find_one({"id": cylinder_id}, {"_id": 0})
    return updated

@api_router.delete("/gas-cylinders/{cylinder_id}")
async def delete_gas_cylinder(cylinder_id: str, admin: dict = Depends(require_admin)):
    """Delete a gas cylinder"""
    result = await db.gas_cylinders.delete_one({"id": cylinder_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Bouteille non trouvée")
    return {"message": "Bouteille supprimée"}

# ==================== MAINTENANCE CONTRACTS ROUTES ====================

@api_router.get("/contracts", response_model=List[dict])
async def get_contracts(current_user: dict = Depends(get_current_user)):
    """Get all maintenance contracts"""
    contracts = await db.contracts.find({}, {"_id": 0}).to_list(1000)
    return contracts

@api_router.get("/contracts/{contract_id}")
async def get_contract(contract_id: str, current_user: dict = Depends(get_current_user)):
    """Get a single contract"""
    contract = await db.contracts.find_one({"id": contract_id}, {"_id": 0})
    if not contract:
        raise HTTPException(status_code=404, detail="Contrat non trouvé")
    return contract

@api_router.post("/contracts", response_model=dict)
async def create_contract(contract: MaintenanceContractCreate, current_user: dict = Depends(require_admin)):
    """Create a new maintenance contract"""
    contract_obj = MaintenanceContract(**contract.model_dump())
    doc = contract_obj.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    await db.contracts.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.put("/contracts/{contract_id}")
async def update_contract(contract_id: str, contract: MaintenanceContractCreate, current_user: dict = Depends(require_admin)):
    """Update a contract"""
    existing = await db.contracts.find_one({"id": contract_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Contrat non trouvé")
    
    update_data = contract.model_dump()
    await db.contracts.update_one({"id": contract_id}, {"$set": update_data})
    updated = await db.contracts.find_one({"id": contract_id}, {"_id": 0})
    return updated

@api_router.delete("/contracts/{contract_id}")
async def delete_contract(contract_id: str, admin: dict = Depends(require_admin)):
    """Delete a contract"""
    result = await db.contracts.delete_one({"id": contract_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Contrat non trouvé")
    return {"message": "Contrat supprimé"}

# ==================== DOCUMENTS (GESTION DOCUMENTAIRE) ROUTES ====================

@api_router.get("/documents", response_model=List[dict])
async def get_documents(
    type_document: Optional[str] = None,
    equipment_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get all documents with optional filters"""
    query = {}
    if type_document:
        query["type_document"] = type_document
    if equipment_id:
        query["equipment_id"] = equipment_id
    
    documents = await db.documents.find(query, {"_id": 0}).to_list(1000)
    return documents

@api_router.post("/documents", response_model=dict)
async def create_document(document: DocumentCreate, current_user: dict = Depends(require_technicien_or_admin)):
    """Create a new document record"""
    document_obj = Document(**document.model_dump())
    doc = document_obj.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    doc["uploaded_by"] = current_user.get("id")
    await db.documents.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.post("/documents/upload")
async def upload_document(
    file: UploadFile = File(...),
    titre: str = Form(...),
    type_document: str = Form(...),
    categorie: Optional[str] = Form(None),
    equipment_id: Optional[str] = Form(None),
    date_validite: Optional[str] = Form(None),
    current_user: dict = Depends(require_technicien_or_admin)
):
    """Upload a document file"""
    file_id = str(uuid.uuid4())
    file_ext = Path(file.filename).suffix
    filename = f"{file_id}{file_ext}"
    file_path = UPLOADS_DIR / "documents" / filename
    
    with open(file_path, "wb") as f:
        shutil.copyfileobj(file.file, f)
    
    document_obj = Document(
        titre=titre,
        type_document=type_document,
        categorie=categorie,
        equipment_id=equipment_id,
        date_validite=date_validite,
        fichier_url=f"/api/uploads/documents/{filename}",
        fichier_nom=file.filename,
        uploaded_by=current_user.get("id")
    )
    
    doc = document_obj.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    await db.documents.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.delete("/documents/{document_id}")
async def delete_document(document_id: str, admin: dict = Depends(require_admin)):
    """Delete a document"""
    doc = await db.documents.find_one({"id": document_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Document non trouvé")
    
    # Delete file if exists
    if doc.get("fichier_url"):
        filename = doc["fichier_url"].split("/")[-1]
        file_path = UPLOADS_DIR / "documents" / filename
        if file_path.exists():
            file_path.unlink()
    
    await db.documents.delete_one({"id": document_id})
    return {"message": "Document supprimé"}

# ==================== BUDGET ROUTES ====================

@api_router.get("/budget", response_model=List[dict])
async def get_budget_items(
    annee: Optional[int] = None,
    categorie: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get budget items with optional filters"""
    query = {}
    if annee:
        query["annee"] = annee
    if categorie:
        query["categorie"] = categorie
    
    items = await db.budget.find(query, {"_id": 0}).to_list(1000)
    return items

@api_router.get("/budget/summary/{annee}")
async def get_budget_summary(annee: int, current_user: dict = Depends(get_current_user)):
    """Get budget summary for a year"""
    items = await db.budget.find({"annee": annee}, {"_id": 0}).to_list(1000)
    
    summary = {
        "annee": annee,
        "total_prevu_xpf": 0,
        "total_prevu_eur": 0,
        "total_realise_xpf": 0,
        "total_realise_eur": 0,
        "par_categorie": {},
        "items": items
    }
    
    categories = {}
    for item in items:
        cat = item.get("categorie", "autre")
        if cat not in categories:
            categories[cat] = {"prevu_xpf": 0, "prevu_eur": 0, "realise_xpf": 0, "realise_eur": 0, "count": 0}
        
        prevu_xpf = item.get("montant_prevu_xpf", 0) or 0
        realise_xpf = item.get("montant_realise_xpf", 0) or 0
        
        categories[cat]["prevu_xpf"] += prevu_xpf
        categories[cat]["prevu_eur"] += prevu_xpf * XPF_TO_EUR
        categories[cat]["realise_xpf"] += realise_xpf
        categories[cat]["realise_eur"] += realise_xpf * XPF_TO_EUR
        categories[cat]["count"] += 1
        
        summary["total_prevu_xpf"] += prevu_xpf
        summary["total_realise_xpf"] += realise_xpf
    
    summary["total_prevu_eur"] = round(summary["total_prevu_xpf"] * XPF_TO_EUR, 2)
    summary["total_realise_eur"] = round(summary["total_realise_xpf"] * XPF_TO_EUR, 2)
    summary["par_categorie"] = categories
    
    return summary

@api_router.post("/budget", response_model=dict)
async def create_budget_item(item: BudgetItemCreate, current_user: dict = Depends(require_admin)):
    """Create a new budget item"""
    item_obj = BudgetItem(**item.model_dump())
    doc = item_obj.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    # Auto-calculate EUR
    doc["montant_prevu_eur"] = round(doc.get("montant_prevu_xpf", 0) * XPF_TO_EUR, 2)
    if doc.get("montant_realise_xpf"):
        doc["montant_realise_eur"] = round(doc["montant_realise_xpf"] * XPF_TO_EUR, 2)
    await db.budget.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.put("/budget/{item_id}")
async def update_budget_item(item_id: str, item: BudgetItemCreate, current_user: dict = Depends(require_admin)):
    """Update a budget item"""
    existing = await db.budget.find_one({"id": item_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Élément budgétaire non trouvé")
    
    update_data = item.model_dump()
    # Auto-calculate EUR
    update_data["montant_prevu_eur"] = round(update_data.get("montant_prevu_xpf", 0) * XPF_TO_EUR, 2)
    if update_data.get("montant_realise_xpf"):
        update_data["montant_realise_eur"] = round(update_data["montant_realise_xpf"] * XPF_TO_EUR, 2)
    
    await db.budget.update_one({"id": item_id}, {"$set": update_data})
    updated = await db.budget.find_one({"id": item_id}, {"_id": 0})
    return updated

@api_router.delete("/budget/{item_id}")
async def delete_budget_item(item_id: str, admin: dict = Depends(require_admin)):
    """Delete a budget item"""
    result = await db.budget.delete_one({"id": item_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Élément budgétaire non trouvé")
    return {"message": "Élément budgétaire supprimé"}

# ==================== REPORT TEMPLATES ROUTES ====================

@api_router.get("/report-templates", response_model=List[dict])
async def get_report_templates(current_user: dict = Depends(get_current_user)):
    """Get all report templates"""
    templates = await db.report_templates.find({}, {"_id": 0}).to_list(1000)
    return templates

@api_router.post("/report-templates", response_model=dict)
async def create_report_template(template: ReportTemplateCreate, current_user: dict = Depends(require_admin)):
    """Create a new report template"""
    template_obj = ReportTemplate(**template.model_dump())
    doc = template_obj.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    await db.report_templates.insert_one(doc)
    doc.pop("_id", None)
    return doc

# ==================== CONTROL REPORTS (PV) ROUTES ====================

@api_router.get("/control-reports", response_model=List[dict])
async def get_control_reports(
    equipment_id: Optional[str] = None,
    template_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get all control reports"""
    query = {}
    if equipment_id:
        query["equipment_id"] = equipment_id
    if template_id:
        query["template_id"] = template_id
    
    reports = await db.control_reports.find(query, {"_id": 0}).to_list(1000)
    return reports

@api_router.post("/control-reports", response_model=dict)
async def create_control_report(report: ControlReportCreate, current_user: dict = Depends(require_technicien_or_admin)):
    """Create a new control report"""
    report_obj = ControlReport(**report.model_dump())
    doc = report_obj.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    await db.control_reports.insert_one(doc)
    doc.pop("_id", None)
    return doc

# ==================== EXCEL IMPORT ROUTES ====================

TEMPLATES = {
    "equipements": {
        "sheet": "Equipements",
        "headers": ["REFERENCE", "TYPE", "MARQUE", "MODELE", "N_SERIE", "DATE_INSTALLATION", "CRITICITE", "STATUT", "LOCALISATION", "COMPTEUR_HORAIRE"],
        "example": ["BAUER 01", "Compresseur", "Bauer", "Mariner 320", "150-11-5_5200-3227", "01/09/2000", "haute", "en_service", "Local compresseurs", "7002"],
    },
    "interventions": {
        "sheet": "Interventions",
        "headers": ["EQUIPEMENT", "N_SERIE", "TYPE", "DATE", "INTERVENANT", "ACTIONS_REALISEES", "OBSERVATION", "COMPTEUR_HORAIRE", "PIECES_UTILISEES"],
        "example": ["BAUER 01", "", "preventive", "15/06/2026", "Radek T.", "Contrôle et appoint d'huile", "RAS", "7002", "Filtre à air x1"],
    },
    "sous-equipements": {
        "sheet": "Sous-equipements",
        "headers": ["PARENT_EQUIPEMENT", "NOM", "REFERENCE", "N_SERIE", "DATE_INSTALLATION", "STATUT", "DESCRIPTION"],
        "example": ["Pupitre de commande", "Manomètre HP CHPF", "MANO-CHPF-01", "", "01/09/2000", "en_service", "Manomètre haute pression (soupape/déverseur/manomètre)"],
    },
    "maintenance": {
        "sheet": "Maintenances",
        "headers": ["EQUIPEMENT", "TITRE", "DESCRIPTION", "PERIODICITE_JOURS", "PERIODICITE_HEURES", "DATE_PLANIFIEE", "PRIORITE", "TECHNICIEN"],
        "example": ["BAUER 01", "Vidange huile compresseur", "Vidange + remplacement des filtres", "365", "1000", "15/09/2026", "haute", "Radek T."],
    },
    "controles": {
        "sheet": "Controles",
        "headers": ["EQUIPEMENT", "TITRE", "TYPE_CONTROLE", "PERIODICITE", "DATE_REALISATION", "ORGANISME", "RESULTAT", "OBSERVATIONS"],
        "example": ["Caisson hyperbare", "Requalification périodique", "reglementaire", "biannuel", "10/03/2025", "APAVE", "conforme", "RAS"],
    },
}


@api_router.get("/import/template/{import_type}")
async def download_import_template(import_type: str, current_user: dict = Depends(get_current_user)):
    """Télécharge un modèle Excel pré-formaté pour l'import."""
    tpl = TEMPLATES.get(import_type)
    if not tpl:
        raise HTTPException(status_code=404, detail="Modèle indisponible pour ce type")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = tpl["sheet"]
    ws.append(tpl["headers"])
    ws.append(tpl["example"])
    header_fill = PatternFill(start_color="005F73", end_color="005F73", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for i, h in enumerate(tpl["headers"], 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(16, len(h) + 4)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"modele_import_{import_type}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )



@api_router.post("/import/excel")
async def import_excel_file(
    file: UploadFile = File(...),
    import_type: str = Form(...),  # maintenance, controles, bouteilles, budget, equipements, interventions
    current_user: dict = Depends(require_admin)
):
    """Import data from Excel/CSV/Numbers file"""
    fname = (file.filename or "").lower()
    ext = None
    for e in (".xlsx", ".xls", ".csv", ".numbers"):
        if fname.endswith(e):
            ext = e
            break
    if not ext:
        raise HTTPException(status_code=400, detail="Format non supporté. Utilisez .xlsx, .csv ou .numbers")

    # Save file temporarily (en conservant l'extension)
    file_id = str(uuid.uuid4())
    temp_path = UPLOADS_DIR / "imports" / f"{file_id}{ext}"

    with open(temp_path, "wb") as f:
        shutil.copyfileobj(file.file, f)

    try:
        result = {"imported": 0, "errors": [], "type": import_type}

        if import_type == "prestataires":
            result = await import_contractors_from_excel(temp_path)
        elif import_type == "bouteilles":
            result = await import_gas_cylinders_from_excel(temp_path)
        elif import_type == "budget":
            result = await import_budget_from_excel(temp_path)
        elif import_type == "maintenance":
            result = await import_maintenance_from_rows(_read_tabular_rows(temp_path, ext))
        elif import_type == "controles":
            result = await import_controls_from_rows(_read_tabular_rows(temp_path, ext))
        elif import_type == "sous-equipements":
            result = await import_subequipments_from_rows(_read_tabular_rows(temp_path, ext))
        elif import_type == "equipements":
            result = await import_equipements_from_rows(_read_tabular_rows(temp_path, ext))
        elif import_type == "interventions":
            result = await import_interventions_from_rows(_read_tabular_rows(temp_path, ext))
        else:
            raise HTTPException(status_code=400, detail=f"Type d'import non supporté: {import_type}")

        return result

    finally:
        # Clean up temp file
        if temp_path.exists():
            temp_path.unlink()


def _read_tabular_rows(file_path: Path, ext: str):
    """Lit un fichier .xlsx/.csv/.numbers -> liste de dict (clés en MAJUSCULES)."""
    if ext == ".numbers":
        try:
            from numbers_parser import Document
        except Exception:
            raise HTTPException(status_code=400, detail="Fichier .numbers non pris en charge sur ce serveur. Exportez-le en Excel (.xlsx) depuis Numbers : Fichier → Exporter vers → Excel.")
        doc = Document(str(file_path))
        raw = doc.sheets[0].tables[0].rows(values_only=True)
        header_idx = next((i for i, r in enumerate(raw) if r and any(c not in (None, "") for c in r)), 0)
        headers = [str(h).strip().upper() if h is not None else "" for h in raw[header_idx]]
        rows = []
        for r in raw[header_idx + 1:]:
            if not r or not any(c not in (None, "") for c in r):
                continue
            rows.append({headers[i]: r[i] for i in range(min(len(headers), len(r))) if headers[i]})
        return rows
    if ext == ".csv":
        df = pd.read_csv(file_path, dtype=str)
    else:
        df = pd.read_excel(file_path, sheet_name=0)
    df.columns = [str(c).strip().upper() for c in df.columns]
    return df.to_dict("records")


def _cell(row: dict, *keys):
    for k in keys:
        if k in row:
            v = row[k]
            if v is None:
                continue
            try:
                if isinstance(v, float) and pd.isna(v):
                    continue
            except Exception:
                pass
            s = str(v).strip()
            if s and s.lower() != "nan":
                return s
    return None


def _parse_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s or s.lower() == "nan":
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s[:10], fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    try:
        return pd.to_datetime(s, dayfirst=True).strftime("%Y-%m-%d")
    except Exception:
        return None


def _to_float(v):
    s = _cell({"X": v}, "X")
    if s is None:
        return None
    try:
        return float(s.replace(",", ".").replace(" ", ""))
    except ValueError:
        return None


async def import_equipements_from_rows(rows: list) -> dict:
    imported, updated, errors = 0, 0, []
    caisson = await db.equipments.find_one({"type": {"$regex": "caisson", "$options": "i"}}, {"_id": 0, "id": 1})
    default_caisson = caisson["id"] if caisson else ""
    existing_types = {t["nom"] async for t in db.equipment_types.find({}, {"_id": 0, "nom": 1})}
    for i, row in enumerate(rows):
        ref = _cell(row, "REFERENCE", "RÉFÉRENCE", "REF")
        typ = _cell(row, "TYPE")
        if not ref or not typ:
            errors.append(f"Ligne {i + 2}: REFERENCE et TYPE obligatoires")
            continue
        if typ not in existing_types:
            await db.equipment_types.insert_one({"id": str(uuid.uuid4()), "nom": typ, "created_at": datetime.now(timezone.utc).isoformat()})
            existing_types.add(typ)
        crit = (_cell(row, "CRITICITE", "CRITICITÉ") or "normale").lower()
        statut = (_cell(row, "STATUT") or "en_service").lower()
        fields = {
            "type": typ,
            "reference": ref,
            "numero_serie": _cell(row, "N_SERIE", "NUMERO_SERIE", "N° SERIE") or ref,
            "marque": _cell(row, "MARQUE"),
            "modele": _cell(row, "MODELE", "MODÈLE"),
            "criticite": crit if crit in ("critique", "haute", "normale", "basse") else "normale",
            "statut": statut if statut in ("en_service", "maintenance", "hors_service", "reforme") else "en_service",
            "date_installation": _parse_date(_cell(row, "DATE_INSTALLATION", "DATE INSTALLATION")),
            "localisation": _cell(row, "LOCALISATION"),
            "compteur_horaire": _to_float(_cell(row, "COMPTEUR_HORAIRE", "COMPTEUR HORAIRE")),
        }
        fields = {k: v for k, v in fields.items() if v is not None}
        existing = await db.equipments.find_one({"reference": ref}, {"_id": 0, "id": 1})
        if existing:
            await db.equipments.update_one({"id": existing["id"]}, {"$set": fields})
            updated += 1
        else:
            fields["id"] = str(uuid.uuid4())
            fields.setdefault("caisson_id", default_caisson)
            fields["historique_statut"] = []
            fields["created_at"] = datetime.now(timezone.utc).isoformat()
            await db.equipments.insert_one(fields)
            imported += 1
    return {"imported": imported, "updated": updated, "errors": errors, "type": "equipements"}


async def import_interventions_from_rows(rows: list) -> dict:
    imported, errors = 0, []
    by_ref, by_serie = {}, {}
    async for e in db.equipments.find({}, {"_id": 0, "id": 1, "reference": 1, "numero_serie": 1}):
        if e.get("reference"):
            by_ref[e["reference"].strip().lower()] = e["id"]
        if e.get("numero_serie"):
            by_serie[str(e["numero_serie"]).strip().lower()] = e["id"]
    # Sous-équipements : garder l'id du sous-équipement ET l'id de l'équipement parent
    sub_map = {}
    async for s in db.subequipments.find({}, {"_id": 0, "id": 1, "reference": 1, "numero_serie": 1, "nom": 1, "parent_equipment_id": 1}):
        for key in (s.get("reference"), s.get("numero_serie"), s.get("nom")):
            if key:
                sub_map.setdefault(str(key).strip().lower(), {"sub_id": s["id"], "parent_id": s.get("parent_equipment_id")})
    docs = []
    for i, row in enumerate(rows):
        eqref = _cell(row, "EQUIPEMENT", "EQUIPEMENT ", "EQUIPMENT")
        serie = _cell(row, "N_SERIE", "NUMERO_SERIE")
        eqid, subid = None, None
        key_lc = (eqref or serie or "").strip().lower()
        if eqref:
            eqid = by_ref.get(eqref.strip().lower())
        if not eqid and serie:
            eqid = by_serie.get(serie.strip().lower())
        # Rattachement à un sous-équipement (par référence/nom/série)
        if not eqid and key_lc in sub_map:
            subid = sub_map[key_lc]["sub_id"]
            eqid = sub_map[key_lc]["parent_id"]
        if not eqid and not subid:
            errors.append(f"Ligne {i + 2}: équipement introuvable (EQUIPEMENT='{eqref}', N_SERIE='{serie}')")
            continue
        date = _parse_date(_cell(row, "DATE"))
        titre = _cell(row, "TITRE", "DESIGNATION", "DÉSIGNATION", "MOTIF")
        actions = _cell(row, "ACTIONS_REALISEES", "ACTIONS REALISEES", "ACTIONS RÉALISÉES", "INTERVENTIONS", "ACTIONS") or titre
        if not date or not actions:
            errors.append(f"Ligne {i + 2}: DATE et ACTIONS_REALISEES (ou DESIGNATION) obligatoires")
            continue
        typ = (_cell(row, "TYPE") or "curative").lower()
        typ = "preventive" if typ.startswith("prev") or typ.startswith("prév") else "curative"
        pieces_txt = _cell(row, "PIECES_UTILISEES", "PIECES UTILISEES", "PIÈCES")
        obs = _cell(row, "OBSERVATION", "OBSERVATIONS")
        if pieces_txt:
            obs = f"{obs} | Pièces: {pieces_txt}" if obs else f"Pièces: {pieces_txt}"
        docs.append({
            "id": str(uuid.uuid4()),
            "work_order_id": None,
            "maintenance_preventive_id": None,
            "type_intervention": typ,
            "titre": titre or actions,
            "date_intervention": date,
            "technicien": _cell(row, "INTERVENANT", "INTERVENANTS", "TECHNICIEN") or "Import",
            "actions_realisees": actions,
            "observations": obs,
            "pieces_utilisees": [],
            "duree_minutes": None,
            "compteur_horaire": _to_float(_cell(row, "COMPTEUR_HORAIRE", "COMPTEUR HORAIRE")),
            "equipment_id": eqid,
            "sous_equipement_id": subid,
            "source": "import_excel",
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
    if docs:
        await db.interventions.insert_many(docs)
        imported = len(docs)
    return {"imported": imported, "errors": errors, "type": "interventions"}


async def import_subequipments_from_rows(rows: list) -> dict:
    """Import des sous-équipements (soupapes, manomètres, déverseurs...) rattachés à un équipement parent."""
    imported, updated, errors = 0, 0, []
    by_ref = {}
    async for e in db.equipments.find({}, {"_id": 0, "id": 1, "reference": 1, "numero_serie": 1}):
        if e.get("reference"):
            by_ref[e["reference"].strip().lower()] = e["id"]
        if e.get("numero_serie"):
            by_ref.setdefault(str(e["numero_serie"]).strip().lower(), e["id"])
    for i, row in enumerate(rows):
        parent = _cell(row, "PARENT_EQUIPEMENT", "PARENT", "EQUIPEMENT")
        nom = _cell(row, "NOM", "DESIGNATION", "DÉSIGNATION")
        if not parent or not nom:
            errors.append(f"Ligne {i + 2}: PARENT_EQUIPEMENT et NOM obligatoires")
            continue
        pid = by_ref.get(parent.strip().lower())
        if not pid:
            errors.append(f"Ligne {i + 2}: équipement parent introuvable ('{parent}')")
            continue
        ref = _cell(row, "REFERENCE", "RÉFÉRENCE", "REF") or nom
        statut = (_cell(row, "STATUT") or "en_service").lower()
        fields = {
            "nom": nom,
            "reference": ref,
            "numero_serie": _cell(row, "N_SERIE", "NUMERO_SERIE"),
            "parent_equipment_id": pid,
            "description": _cell(row, "DESCRIPTION", "TYPE"),
            "date_installation": _parse_date(_cell(row, "DATE_INSTALLATION", "DATE INSTALLATION")),
            "statut": statut if statut in ("en_service", "maintenance", "hors_service") else "en_service",
        }
        clean = {k: v for k, v in fields.items() if v is not None}
        existing = await db.subequipments.find_one({"reference": ref, "parent_equipment_id": pid}, {"_id": 0, "id": 1})
        if existing:
            await db.subequipments.update_one({"id": existing["id"]}, {"$set": clean})
            updated += 1
        else:
            clean["id"] = str(uuid.uuid4())
            clean["photos"] = []
            clean["documents"] = []
            clean["created_at"] = datetime.now(timezone.utc).isoformat()
            await db.subequipments.insert_one(clean)
            imported += 1
    return {"imported": imported, "updated": updated, "errors": errors, "type": "sous-equipements"}


async def import_maintenance_from_rows(rows: list) -> dict:
    """Import des maintenances préventives (ordres de travail préventifs récurrents)."""
    imported, errors = 0, []
    by_ref = {}
    async for e in db.equipments.find({}, {"_id": 0, "id": 1, "reference": 1, "numero_serie": 1, "caisson_id": 1}):
        if e.get("reference"):
            by_ref[e["reference"].strip().lower()] = e
        if e.get("numero_serie"):
            by_ref.setdefault(str(e["numero_serie"]).strip().lower(), e)
    docs = []
    for i, row in enumerate(rows):
        titre = _cell(row, "TITRE", "INTITULE", "DESIGNATION", "DETAIL INTERVENTIONS")
        if not titre:
            errors.append(f"Ligne {i + 2}: TITRE obligatoire")
            continue
        eqref = _cell(row, "EQUIPEMENT", "EQUIPMENT")
        eq = by_ref.get(eqref.strip().lower()) if eqref else None
        if eqref and not eq:
            errors.append(f"Ligne {i + 2}: équipement introuvable ('{eqref}')")
            continue
        prio = (_cell(row, "PRIORITE", "PRIORITÉ") or "normale").lower()
        pj = _to_float(_cell(row, "PERIODICITE_JOURS", "PÉRIODICITÉ_JOURS", "PERIODICITE JOURS"))
        ph = _to_float(_cell(row, "PERIODICITE_HEURES", "PÉRIODICITÉ_HEURES", "PERIODICITE HEURES"))
        docs.append({
            "id": str(uuid.uuid4()),
            "titre": titre,
            "description": _cell(row, "DESCRIPTION", "OBSERVATION") or titre,
            "type_maintenance": "preventive",
            "priorite": prio if prio in ("urgente", "haute", "normale", "basse") else "normale",
            "statut": "planifiee",
            "caisson_id": eq.get("caisson_id") if eq else None,
            "equipment_id": eq["id"] if eq else None,
            "date_planifiee": _parse_date(_cell(row, "DATE_PLANIFIEE", "DATE PLANIFIEE", "DATE")) or datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            "periodicite_jours": int(pj) if pj else None,
            "periodicite_heures": int(ph) if ph else None,
            "technicien_assigne": _cell(row, "TECHNICIEN", "INTERVENANT"),
            "photos": [],
            "documents": [],
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
    updated = 0
    for doc in docs:
        # Anti-duplication: match by (equipment_id, titre)
        existing = await db.work_orders.find_one({
            "equipment_id": doc["equipment_id"],
            "titre": doc["titre"],
            "type_maintenance": "preventive",
        })
        if existing:
            upd = {k: v for k, v in doc.items() if k not in ("id", "created_at", "photos", "documents")}
            await db.work_orders.update_one({"id": existing["id"]}, {"$set": upd})
            updated += 1
        else:
            await db.work_orders.insert_one(doc)
            imported += 1
    return {"imported": imported, "updated": updated, "errors": errors, "type": "maintenance"}


async def import_controls_from_rows(rows: list) -> dict:
    """Import des contrôles périodiques réglementaires (inspections)."""
    imported, errors = 0, []
    by_ref = {}
    async for e in db.equipments.find({}, {"_id": 0, "id": 1, "reference": 1, "numero_serie": 1, "caisson_id": 1}):
        if e.get("reference"):
            by_ref[e["reference"].strip().lower()] = e
        if e.get("numero_serie"):
            by_ref.setdefault(str(e["numero_serie"]).strip().lower(), e)
    docs = []
    for i, row in enumerate(rows):
        titre = _cell(row, "TITRE", "CONTROLE", "CONTRÔLE", "DESIGNATION")
        if not titre:
            errors.append(f"Ligne {i + 2}: TITRE obligatoire")
            continue
        eqref = _cell(row, "EQUIPEMENT", "EQUIPMENT")
        eq = by_ref.get(eqref.strip().lower()) if eqref else None
        if eqref and not eq:
            errors.append(f"Ligne {i + 2}: équipement introuvable ('{eqref}')")
            continue
        per = _norm_periodicite(_cell(row, "PERIODICITE", "PÉRIODICITÉ", "PERIODICITE_JOURS"))
        date_real = _parse_date(_cell(row, "DATE_REALISATION", "DATE REALISATION", "DATE"))
        date_val = None
        if date_real:
            try:
                date_val = (datetime.strptime(date_real, "%Y-%m-%d") + timedelta(days=PERIODICITES[per])).strftime("%Y-%m-%d")
            except Exception:
                pass
        docs.append({
            "id": str(uuid.uuid4()),
            "titre": titre,
            "type_controle": _cell(row, "TYPE_CONTROLE", "TYPE", "DESCRIPTION") or "reglementaire",
            "periodicite": per,
            "caisson_id": eq.get("caisson_id") if eq else None,
            "equipment_id": eq["id"] if eq else None,
            "date_realisation": date_real,
            "date_validite": date_val,
            "organisme_certificateur": _cell(row, "ORGANISME", "ORGANISME_CERTIFICATEUR"),
            "resultat": _cell(row, "RESULTAT", "RÉSULTAT"),
            "observations": _cell(row, "OBSERVATIONS", "OBSERVATION"),
            "procedure_documents": [],
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
    updated = 0
    for doc in docs:
        # Anti-duplication: match by (equipment_id, titre)
        existing = await db.inspections.find_one({
            "equipment_id": doc["equipment_id"],
            "titre": doc["titre"],
        })
        if existing:
            upd = {k: v for k, v in doc.items() if k not in ("id", "created_at", "procedure_documents")}
            await db.inspections.update_one({"id": existing["id"]}, {"$set": upd})
            updated += 1
        else:
            await db.inspections.insert_one(doc)
            imported += 1
    return {"imported": imported, "updated": updated, "errors": errors, "type": "controles"}



async def import_contractors_from_excel(file_path: Path) -> dict:
    """Import contractors from Excel"""
    df = pd.read_excel(file_path, sheet_name=0)
    imported = 0
    errors = []
    
    # Default contractors to create
    default_contractors = [
        {"nom": "Bauer Nautisport", "type": "prestataire", "specialite": "Maintenance compresseurs"},
        {"nom": "Comex", "type": "prestataire", "specialite": "Maintenance caisson hyperbare"},
        {"nom": "Métrologie de Tahiti", "type": "prestataire", "specialite": "Étalonnage, métrologie"},
        {"nom": "Gazpac", "type": "fournisseur", "specialite": "Bouteilles de gaz, requalification"},
        {"nom": "BCP", "type": "organisme_controle", "specialite": "Contrôles réglementaires"},
        {"nom": "Argos", "type": "prestataire", "specialite": "ARI, équipements respiratoires"},
        {"nom": "Incendie Moz", "type": "prestataire", "specialite": "Extincteurs"},
        {"nom": "Servomex", "type": "fournisseur", "specialite": "Analyseurs de gaz"},
        {"nom": "Bureau Véritas", "type": "organisme_controle", "specialite": "Contrôles réglementaires"},
        {"nom": "Protais/Vigil", "type": "fournisseur", "specialite": "Manomètres"},
        {"nom": "Nuova Fima", "type": "fournisseur", "specialite": "Manomètres"},
        {"nom": "H+ Valves", "type": "fournisseur", "specialite": "Soupapes"},
        {"nom": "RS Components SAS", "type": "fournisseur", "specialite": "Composants électroniques"},
        {"nom": "FIT", "type": "prestataire", "specialite": "Maintenance générale"},
        {"nom": "Vinci", "type": "prestataire", "specialite": "Filtres"},
    ]
    
    for contractor_data in default_contractors:
        existing = await db.contractors.find_one({"nom": contractor_data["nom"]})
        if not existing:
            contractor_obj = Contractor(**contractor_data)
            doc = contractor_obj.model_dump()
            doc["created_at"] = doc["created_at"].isoformat()
            await db.contractors.insert_one(doc)
            imported += 1
    
    return {"imported": imported, "errors": errors, "type": "prestataires"}

async def import_gas_cylinders_from_excel(file_path: Path) -> dict:
    """Import gas cylinders from Excel"""
    imported = 0
    errors = []
    
    try:
        df = pd.read_excel(file_path, sheet_name=0)
        
        # Map column names (adapt based on actual Excel structure)
        for idx, row in df.iterrows():
            try:
                type_gaz = str(row.get('Nature du gaz', row.get('type_gaz', ''))).strip().lower()
                
                # Normalize gas type
                if 'o2' in type_gaz or 'oxygène' in type_gaz or 'oxygene' in type_gaz:
                    type_gaz = 'O2'
                elif 'air' in type_gaz and 'méd' in type_gaz:
                    type_gaz = 'air_medicale'
                elif 'héliox' in type_gaz or 'heliox' in type_gaz:
                    type_gaz = 'heliox'
                elif 'nitrox' in type_gaz:
                    type_gaz = 'nitrox'
                else:
                    continue  # Skip unknown gas types
                
                numero = str(row.get('N° de la bout.', row.get('numero_bouteille', idx))).strip()
                if not numero or numero == 'nan':
                    numero = f"B-{idx}"
                
                volume = str(row.get('Vol. de bout.', 'B50')).strip()
                if volume == 'nan':
                    volume = 'B50'
                
                cylinder_data = {
                    "numero_bouteille": numero,
                    "type_gaz": type_gaz,
                    "volume": volume,
                    "agent_responsable": str(row.get('Nom de l\'agent', '')).strip() if pd.notna(row.get('Nom de l\'agent')) else None,
                    "observations": str(row.get('Etat / Observations', '')).strip() if pd.notna(row.get('Etat / Observations')) else None,
                    "statut": "pleine"
                }
                
                existing = await db.gas_cylinders.find_one({"numero_bouteille": numero, "type_gaz": type_gaz})
                if not existing:
                    cylinder_obj = GasCylinder(**cylinder_data)
                    doc = cylinder_obj.model_dump()
                    doc["created_at"] = doc["created_at"].isoformat()
                    await db.gas_cylinders.insert_one(doc)
                    imported += 1
                    
            except Exception as e:
                errors.append(f"Ligne {idx}: {str(e)}")
    
    except Exception as e:
        errors.append(f"Erreur de lecture du fichier: {str(e)}")
    
    return {"imported": imported, "errors": errors, "type": "bouteilles"}

async def import_budget_from_excel(file_path: Path) -> dict:
    """Import budget items from Excel"""
    imported = 0
    errors = []
    
    try:
        df = pd.read_excel(file_path, sheet_name=0)
        
        for idx, row in df.iterrows():
            try:
                designation = str(row.get('DETAIL INTERVENTIONS', row.get('designation', ''))).strip()
                if not designation or designation == 'nan':
                    continue
                
                montant = row.get('Montant (2026)', row.get('montant_prevu_xpf', 0))
                if pd.isna(montant):
                    montant = 0
                else:
                    montant = float(montant)
                
                # Determine category
                categorie = "maintenance_preventive"
                designation_lower = designation.lower()
                if 'remplacer' in designation_lower or 'remplacement' in designation_lower:
                    categorie = "pieces_detachees"
                elif 'contrôle' in designation_lower or 'controle' in designation_lower or 'inspection' in designation_lower:
                    categorie = "controle_reglementaire"
                elif 'étalonnage' in designation_lower or 'etalonnage' in designation_lower:
                    categorie = "controle_reglementaire"
                elif 'entretien' in designation_lower:
                    categorie = "prestation_externe"
                
                fournisseur = str(row.get('Fournisseur', '')).strip()
                
                budget_data = {
                    "annee": 2026,
                    "categorie": categorie,
                    "designation": designation,
                    "montant_prevu_xpf": montant,
                    "montant_prevu_eur": round(montant * XPF_TO_EUR, 2),
                    "periodicite": str(row.get('Fréquence', row.get('periodicite', ''))).strip() if pd.notna(row.get('Fréquence', row.get('periodicite'))) else None,
                    "notes": f"Fournisseur: {fournisseur}" if fournisseur and fournisseur != 'nan' else None
                }
                
                item_obj = BudgetItem(**budget_data)
                doc = item_obj.model_dump()
                doc["created_at"] = doc["created_at"].isoformat()
                await db.budget.insert_one(doc)
                imported += 1
                
            except Exception as e:
                errors.append(f"Ligne {idx}: {str(e)}")
    
    except Exception as e:
        errors.append(f"Erreur de lecture du fichier: {str(e)}")
    
    return {"imported": imported, "errors": errors, "type": "budget"}

async def import_maintenance_from_excel(file_path: Path) -> dict:
    """Import maintenance records from Excel"""
    imported = 0
    errors = []
    
    try:
        df = pd.read_excel(file_path, sheet_name=0)
        
        for idx, row in df.iterrows():
            try:
                detail = str(row.get('DETAIL INTERVENTIONS', '')).strip()
                if not detail or detail == 'nan':
                    continue
                
                # Create inspection record for tracking
                inspection_data = {
                    "titre": detail,
                    "type_controle": "maintenance",
                    "periodicite": str(row.get('Périodicité', 'annuel')).strip() if pd.notna(row.get('Périodicité')) else "annuel",
                    "observations": str(row.get('OBSERVATION TECHNIQUE', '')).strip() if pd.notna(row.get('OBSERVATION TECHNIQUE')) else None
                }
                
                inspection_obj = Inspection(**inspection_data)
                doc = inspection_obj.model_dump()
                doc["created_at"] = doc["created_at"].isoformat()
                await db.inspections.insert_one(doc)
                imported += 1
                
            except Exception as e:
                errors.append(f"Ligne {idx}: {str(e)}")
    
    except Exception as e:
        errors.append(f"Erreur de lecture du fichier: {str(e)}")
    
    return {"imported": imported, "errors": errors, "type": "maintenance"}

async def import_controls_from_excel(file_path: Path) -> dict:
    """Import control records from Excel"""
    imported = 0
    errors = []
    
    # This will be populated with data from the controls Excel file
    return {"imported": imported, "errors": errors, "type": "controles"}

# ==================== INITIALIZE DEFAULT DATA ====================

@api_router.post("/init/default-data")
async def initialize_default_data(admin: dict = Depends(require_admin)):
    """Initialize default prestataires and report templates"""
    results = {"contractors": 0, "templates": 0}
    
    # Create default contractors
    default_contractors = [
        {"nom": "Bauer Nautisport", "type": "prestataire", "specialite": "Maintenance compresseurs"},
        {"nom": "Comex", "type": "prestataire", "specialite": "Maintenance caisson hyperbare"},
        {"nom": "Métrologie de Tahiti", "type": "prestataire", "specialite": "Étalonnage, métrologie"},
        {"nom": "Gazpac", "type": "fournisseur", "specialite": "Bouteilles de gaz, requalification"},
        {"nom": "BCP", "type": "organisme_controle", "specialite": "Contrôles réglementaires"},
        {"nom": "Argos", "type": "prestataire", "specialite": "ARI, équipements respiratoires"},
        {"nom": "Incendie Moz", "type": "prestataire", "specialite": "Extincteurs"},
        {"nom": "Servomex", "type": "fournisseur", "specialite": "Analyseurs de gaz"},
        {"nom": "Bureau Véritas", "type": "organisme_controle", "specialite": "Contrôles réglementaires"},
        {"nom": "Protais/Vigil", "type": "fournisseur", "specialite": "Manomètres"},
        {"nom": "Nuova Fima", "type": "fournisseur", "specialite": "Manomètres"},
        {"nom": "H+ Valves", "type": "fournisseur", "specialite": "Soupapes"},
    ]
    
    for contractor_data in default_contractors:
        existing = await db.contractors.find_one({"nom": contractor_data["nom"]})
        if not existing:
            contractor_obj = Contractor(**contractor_data)
            doc = contractor_obj.model_dump()
            doc["created_at"] = doc["created_at"].isoformat()
            await db.contractors.insert_one(doc)
            results["contractors"] += 1
    
    # Create default report templates
    default_templates = [
        {
            "nom": "Analyse de l'air respirable",
            "type_controle": "analyse_air",
            "description": "Analyse de la qualité de l'air respirable des compresseurs BAUER",
            "champs": [
                {"nom": "marque_compresseur", "type": "text", "obligatoire": True},
                {"nom": "modele", "type": "text", "obligatoire": True},
                {"nom": "numero_serie", "type": "text", "obligatoire": False},
                {"nom": "compteur_horaire", "type": "number", "obligatoire": True},
                {"nom": "h2o_valeur", "type": "number", "obligatoire": True, "unite": "mg/m³"},
                {"nom": "co_valeur", "type": "number", "obligatoire": True, "unite": "ppm"},
                {"nom": "co2_valeur", "type": "number", "obligatoire": True, "unite": "ppm"},
                {"nom": "huile_valeur", "type": "number", "obligatoire": True, "unite": "mg/m³"},
                {"nom": "odeur_gout", "type": "select", "options": ["Aucun", "Léger", "Significatif"], "obligatoire": True}
            ],
            "normes_reference": ["Article N°3 de la délibération N° 87-79 AT du 12 juin 1987"],
            "criteres_conformite": [
                {"parametre": "H2O", "valeur_max": 5, "unite": "mg/m³", "note": "Peut dépasser selon conditions climatiques PF"},
                {"parametre": "CO", "valeur_max": 15, "unite": "ppm"},
                {"parametre": "CO2", "valeur_max": 500, "unite": "ppm"},
                {"parametre": "Huile", "valeur_max": 0.5, "unite": "mg/m³"}
            ]
        },
        {
            "nom": "Contrôle annuel du caisson",
            "type_controle": "controle_annuel",
            "description": "Contrôle annuel complet du caisson hyperbare",
            "champs": [
                {"nom": "etancheite_enceintes", "type": "checkbox", "obligatoire": True},
                {"nom": "etancheite_portes", "type": "checkbox", "obligatoire": True},
                {"nom": "securite_sas_medicament", "type": "checkbox", "obligatoire": True},
                {"nom": "soupapes_surete", "type": "checkbox", "obligatoire": True},
                {"nom": "manometres", "type": "checkbox", "obligatoire": True},
                {"nom": "hublots", "type": "checkbox", "obligatoire": True},
                {"nom": "eclairage", "type": "checkbox", "obligatoire": True},
                {"nom": "communications", "type": "checkbox", "obligatoire": True},
                {"nom": "systeme_incendie", "type": "checkbox", "obligatoire": True},
                {"nom": "circuits_electriques", "type": "checkbox", "obligatoire": True}
            ],
            "normes_reference": ["Arrêté du 15 mars 2000", "Arrêté du 20 novembre 2017"]
        },
        {
            "nom": "Étalonnage manomètre",
            "type_controle": "etalonnage_manometre",
            "description": "Certificat d'étalonnage des manomètres",
            "champs": [
                {"nom": "marque", "type": "text", "obligatoire": True},
                {"nom": "modele", "type": "text", "obligatoire": True},
                {"nom": "numero_serie", "type": "text", "obligatoire": True},
                {"nom": "plage_mesure", "type": "text", "obligatoire": True},
                {"nom": "classe", "type": "text", "obligatoire": False},
                {"nom": "localisation", "type": "text", "obligatoire": True},
                {"nom": "valeur_mesure_1", "type": "number", "obligatoire": True},
                {"nom": "valeur_reference_1", "type": "number", "obligatoire": True},
                {"nom": "ecart_1", "type": "number", "obligatoire": True}
            ]
        },
        {
            "nom": "Contrôle soupape de sûreté",
            "type_controle": "etalonnage_soupape",
            "description": "Contrôle et étalonnage des soupapes de surpression",
            "champs": [
                {"nom": "marque", "type": "text", "obligatoire": True},
                {"nom": "modele", "type": "text", "obligatoire": True},
                {"nom": "numero_serie", "type": "text", "obligatoire": True},
                {"nom": "pression_tarage", "type": "number", "obligatoire": True, "unite": "bar"},
                {"nom": "pression_ouverture", "type": "number", "obligatoire": True, "unite": "bar"},
                {"nom": "localisation", "type": "text", "obligatoire": True}
            ]
        }
    ]
    
    for template_data in default_templates:
        existing = await db.report_templates.find_one({"nom": template_data["nom"]})
        if not existing:
            template_obj = ReportTemplate(**template_data)
            doc = template_obj.model_dump()
            doc["created_at"] = doc["created_at"].isoformat()
            await db.report_templates.insert_one(doc)
            results["templates"] += 1
    
    return {"message": "Données par défaut initialisées", "results": results}

# Include router and middleware
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
